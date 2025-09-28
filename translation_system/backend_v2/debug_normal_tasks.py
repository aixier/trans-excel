#!/usr/bin/env python3
"""Debug script to understand why normal_tasks is 14"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from utils.color_detector import is_yellow_color, is_blue_color

def analyze_normal_tasks():
    file_path = '/mnt/d/work/trans_excel/sfdaf.xlsx'

    # Load workbook for colors
    workbook = load_workbook(file_path, data_only=True)

    # Load with pandas for data
    sheet_name = workbook.sheetnames[0]
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=0)

    print(f"Sheet: {sheet_name}")
    print(f"Total rows: {len(df)}")
    print(f"Columns: {list(df.columns)}")
    print()

    # Get the actual sheet
    sheet = workbook[sheet_name]

    # Identify source and target columns
    columns = list(df.columns)
    source_cols = []
    target_cols = []

    for idx, col in enumerate(columns):
        col_upper = str(col).upper()
        if col_upper in ['CH', 'CN', '中文', 'EN', 'ENGLISH', '英文']:
            source_cols.append(idx)
        elif col_upper in ['TR', 'TH', 'PT', 'VN', 'VI', 'ES', 'IND', 'ID']:
            target_cols.append(idx)

    print(f"Source columns: {[columns[i] for i in source_cols]} (indices: {source_cols})")
    print(f"Target columns: {[columns[i] for i in target_cols]} (indices: {target_cols})")
    print()

    # First pass: identify rows with yellow/blue markings
    rows_with_yellow = set()
    rows_with_blue = set()
    yellow_details = []
    blue_details = []

    for row_idx in range(len(df)):
        for col_idx in range(len(columns)):
            cell_value = df.iloc[row_idx, col_idx]
            if pd.notna(cell_value) and isinstance(cell_value, str) and len(cell_value.strip()) > 0:
                # Get cell color
                cell = sheet.cell(row=row_idx+2, column=col_idx+1)
                if cell.fill and hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
                    color_value = cell.fill.fgColor.rgb
                    if color_value and isinstance(color_value, str):
                        cell_color = f"#{color_value[2:]}" if len(color_value) == 8 else f"#{color_value}"

                        if is_yellow_color(cell_color):
                            rows_with_yellow.add(row_idx)
                            yellow_details.append(f"Row {row_idx+2} (index {row_idx}), Col {columns[col_idx]}")
                        elif is_blue_color(cell_color):
                            rows_with_blue.add(row_idx)
                            blue_details.append(f"Row {row_idx+2} (index {row_idx}), Col {columns[col_idx]}")

    print(f"Rows with yellow cells: {len(rows_with_yellow)} rows")
    print(f"Rows with blue cells: {len(rows_with_blue)} rows")
    print()

    # Second pass: count normal translation tasks (excluding yellow/blue rows)
    normal_tasks = 0
    normal_task_details = []

    for row_idx in range(len(df)):
        # Skip rows that have yellow or blue markings
        if row_idx in rows_with_yellow or row_idx in rows_with_blue:
            continue

        # Check if source columns have content
        has_source = False
        source_text = None
        source_col_name = None

        for source_col_idx in source_cols:
            if source_col_idx < len(columns):
                source_value = df.iloc[row_idx, source_col_idx]
                if pd.notna(source_value) and isinstance(source_value, str) and len(source_value.strip()) > 0:
                    has_source = True
                    source_text = str(source_value).strip()
                    source_col_name = columns[source_col_idx]
                    break

        if has_source and source_text:
            # Count empty target columns
            for target_col_idx in target_cols:
                if target_col_idx < len(columns):
                    target_value = df.iloc[row_idx, target_col_idx]
                    if pd.isna(target_value) or str(target_value).strip() == '':
                        normal_tasks += 1
                        if len(normal_task_details) < 20:  # Show first 20
                            normal_task_details.append(
                                f"Row {row_idx+2}: {source_col_name}='{source_text[:30]}...' -> {columns[target_col_idx]} (empty)"
                            )

    print("=== Normal Tasks Details ===")
    print(f"Total normal tasks: {normal_tasks}")
    print()
    print("Sample normal tasks:")
    for detail in normal_task_details:
        print(f"  {detail}")

    # Analyze which rows are contributing to normal tasks
    print()
    print("=== Row Analysis ===")

    # Count rows with non-empty source but not yellow/blue
    eligible_rows = 0
    for row_idx in range(len(df)):
        if row_idx in rows_with_yellow or row_idx in rows_with_blue:
            continue

        # Check EN column (index 2)
        en_value = df.iloc[row_idx, 2] if 2 < len(columns) else None
        ch_value = df.iloc[row_idx, 1] if 1 < len(columns) else None

        if (pd.notna(en_value) and str(en_value).strip()) or (pd.notna(ch_value) and str(ch_value).strip()):
            eligible_rows += 1

    print(f"Total rows: {len(df)}")
    print(f"Rows with yellow: {len(rows_with_yellow)}")
    print(f"Rows with blue: {len(rows_with_blue)}")
    print(f"Rows eligible for normal tasks: {eligible_rows}")
    print(f"Target columns: {len(target_cols)}")
    print(f"Expected normal tasks: {eligible_rows} × {len(target_cols)} = {eligible_rows * len(target_cols)}")

if __name__ == '__main__':
    analyze_normal_tasks()
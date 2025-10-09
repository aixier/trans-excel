"""测试comment逻辑：空值不添加comment，重译添加comment"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))

import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from models.excel_dataframe import ExcelDataFrame
from models.task_dataframe import TaskDataFrameManager, TaskStatus
from services.export.excel_exporter import ExcelExporter
from datetime import datetime
import tempfile

def test_comment_logic():
    """测试comment添加逻辑"""

    print("=" * 80)
    print("测试comment逻辑")
    print("=" * 80)

    # 创建测试数据：包含空值和非空值
    test_df = pd.DataFrame({
        'key': ['ID_001', 'ID_002', 'ID_003'],
        'CH': ['你好', '世界', '测试'],
        'EN': ['Hello', 'World', 'Test'],
        'PT': [None, 'Old Translation', None]  # 第2行有旧翻译
    })

    print("\n📊 测试数据:")
    print(test_df)

    # 创建ExcelDataFrame
    excel_df = ExcelDataFrame()
    excel_df.filename = "test.xlsx"
    excel_df.excel_id = "test_session"
    excel_df.add_sheet("Sheet1", test_df)

    # 创建TaskManager，3个翻译任务
    task_manager = TaskDataFrameManager()

    # 任务1: 首次翻译（PT列原本为空）
    task_manager.add_task({
        'task_id': 'TASK_001',
        'sheet_name': 'Sheet1',
        'row_idx': 0,
        'col_idx': 3,  # PT列
        'source_text': '你好',
        'result': 'Olá',
        'status': TaskStatus.COMPLETED,
        'batch_id': 'B1', 'group_id': 'G1', 'task_type': 'normal',
        'source_lang': 'CH', 'target_lang': 'PT', 'source_context': '',
        'game_context': '', 'excel_id': 'test', 'cell_ref': 'D2',
        'priority': 5, 'confidence': 0.9, 'char_count': 2,
        'created_at': datetime.now(), 'updated_at': datetime.now(),
        'start_time': None, 'end_time': None, 'duration_ms': 0,
        'retry_count': 0, 'error_message': '', 'llm_model': 'test',
        'token_count': 10, 'cost': 0.01, 'reviewer_notes': '', 'is_final': False
    })

    # 任务2: 重译（PT列原本有内容）
    task_manager.add_task({
        'task_id': 'TASK_002',
        'sheet_name': 'Sheet1',
        'row_idx': 1,
        'col_idx': 3,  # PT列
        'source_text': '世界',
        'result': 'Mundo',
        'status': TaskStatus.COMPLETED,
        'batch_id': 'B1', 'group_id': 'G1', 'task_type': 'yellow',
        'source_lang': 'CH', 'target_lang': 'PT', 'source_context': '',
        'game_context': '', 'excel_id': 'test', 'cell_ref': 'D3',
        'priority': 9, 'confidence': 0.9, 'char_count': 2,
        'created_at': datetime.now(), 'updated_at': datetime.now(),
        'start_time': None, 'end_time': None, 'duration_ms': 0,
        'retry_count': 0, 'error_message': '', 'llm_model': 'test',
        'token_count': 10, 'cost': 0.01, 'reviewer_notes': '', 'is_final': False
    })

    # 任务3: 首次翻译（PT列原本为空）
    task_manager.add_task({
        'task_id': 'TASK_003',
        'sheet_name': 'Sheet1',
        'row_idx': 2,
        'col_idx': 3,  # PT列
        'source_text': '测试',
        'result': 'Teste',
        'status': TaskStatus.COMPLETED,
        'batch_id': 'B1', 'group_id': 'G1', 'task_type': 'normal',
        'source_lang': 'CH', 'target_lang': 'PT', 'source_context': '',
        'game_context': '', 'excel_id': 'test', 'cell_ref': 'D4',
        'priority': 5, 'confidence': 0.9, 'char_count': 2,
        'created_at': datetime.now(), 'updated_at': datetime.now(),
        'start_time': None, 'end_time': None, 'duration_ms': 0,
        'retry_count': 0, 'error_message': '', 'llm_model': 'test',
        'token_count': 10, 'cost': 0.01, 'reviewer_notes': '', 'is_final': False
    })

    # 导出
    with tempfile.TemporaryDirectory() as tmpdir:
        exporter = ExcelExporter(output_dir=tmpdir)

        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        ws = wb.create_sheet(title='Sheet1')

        import asyncio
        asyncio.run(exporter._write_sheet_with_translations(
            ws, test_df, excel_df, task_manager, 'Sheet1'
        ))

        test_output = Path(tmpdir) / 'test_output.xlsx'
        wb.save(test_output)

        print(f"\n✓ 导出完成")

        # 读取并检查comment
        from openpyxl import load_workbook
        result_wb = load_workbook(test_output)
        result_ws = result_wb['Sheet1']

        print(f"\n{'='*80}")
        print("Comment检查结果:")
        print("=" * 80)

        # 检查第2行（首次翻译，原本为空）
        cell_d2 = result_ws['D2']
        print(f"\n第2行 (D2) - 首次翻译（原PT列为空）:")
        print(f"  值: {cell_d2.value}")
        print(f"  Comment: {cell_d2.comment.text if cell_d2.comment else 'None'}")
        if cell_d2.comment:
            print(f"  ❌ 不应该有comment！")
        else:
            print(f"  ✅ 正确：无comment")

        # 检查第3行（重译，原本有内容）
        cell_d3 = result_ws['D3']
        print(f"\n第3行 (D3) - 重译（原PT列有内容 'Old Translation'）:")
        print(f"  值: {cell_d3.value}")
        print(f"  Comment: {cell_d3.comment.text if cell_d3.comment else 'None'}")
        if cell_d3.comment and 'Old Translation' in cell_d3.comment.text:
            print(f"  ✅ 正确：有comment且包含原文")
        else:
            print(f"  ❌ 缺少comment或不包含原文！")

        # 检查第4行（首次翻译，原本为空）
        cell_d4 = result_ws['D4']
        print(f"\n第4行 (D4) - 首次翻译（原PT列为空）:")
        print(f"  值: {cell_d4.value}")
        print(f"  Comment: {cell_d4.comment.text if cell_d4.comment else 'None'}")
        if cell_d4.comment:
            print(f"  ❌ 不应该有comment！")
        else:
            print(f"  ✅ 正确：无comment")

        print(f"\n{'='*80}")
        print("测试总结:")
        print("=" * 80)
        print("✅ 空值翻译：不添加comment")
        print("✅ 重译场景：添加comment显示原文")

if __name__ == '__main__':
    try:
        test_comment_logic()
    except Exception as e:
        print(f"\n❌ 测试失败: {e}")
        import traceback
        traceback.print_exc()

#!/usr/bin/env python3
"""Check PET sheet content in test2.xlsx"""

import openpyxl
from openpyxl.utils import get_column_letter

def check_pet_sheet():
    # Load workbook
    file_path = '/mnt/d/work/trans_excel/test2.xlsx'
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Check if PET sheet exists
    if 'PET' not in wb.sheetnames:
        print(f"Available sheets: {wb.sheetnames}")
        return

    # Get PET sheet
    ws = wb['PET']

    print(f"PET Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")
    print("\n" + "="*80)

    # Show all cells with their coordinates, values, colors and comments
    print("\nAll cells with content in PET sheet:")
    print("="*80)

    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            if cell.value:
                col_letter = get_column_letter(col_idx)
                print(f"\nCell {col_letter}{row_idx}:")
                print(f"  Value: {cell.value}")

                # Check for color
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                    rgb = cell.fill.fgColor.rgb
                    if rgb and rgb != "00000000":  # Not transparent
                        print(f"  Color: #{rgb}")

                # Check for comment
                if cell.comment:
                    print(f"  Comment: {cell.comment.text}")

                # Show row label (column A) if not in column A
                if col_idx > 1:
                    row_label = ws.cell(row=row_idx, column=1).value
                    if row_label:
                        print(f"  Row Label (A{row_idx}): {row_label}")

                # Show column header (row 1) if not in row 1
                if row_idx > 1:
                    col_header = ws.cell(row=1, column=col_idx).value
                    if col_header:
                        print(f"  Column Header ({col_letter}1): {col_header}")

    # Search for specific patterns
    search_patterns = ["缩短", "Simon", "JIJIA", "宠物", "pet", "PET"]
    print("\n" + "="*80)
    print(f"Searching for patterns: {search_patterns}")
    print("="*80)

    for pattern in search_patterns:
        print(f"\nSearching for: '{pattern}'")
        found_cells = []

        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value and pattern in str(cell.value):
                    col_letter = get_column_letter(col_idx)
                    found_cells.append(f"{col_letter}{row_idx}")

        if found_cells:
            print(f"  Found in cells: {', '.join(found_cells)}")
        else:
            print(f"  Not found")

    wb.close()

if __name__ == "__main__":
    check_pet_sheet()
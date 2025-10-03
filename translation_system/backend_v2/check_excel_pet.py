#!/usr/bin/env python3
"""Check PET sheet content in test2.xlsx"""

import openpyxl
from openpyxl.utils import get_column_letter

def check_pet_sheet():
    # Load workbook
    file_path = '/mnt/d/work/trans_excel/test2.xlsx'
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Check if PET sheet exists
    if 'PET' not in wb.sheetnames:
        print(f"Available sheets: {wb.sheetnames}")
        return

    # Get PET sheet
    ws = wb['PET']

    print(f"PET Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")
    print("\n" + "="*80)

    # Show all content and search for specific patterns
    print("\nShowing all PET sheet content:")
    print("="*80)

    # First, show headers
    print("\nColumn Headers (Row 1):")
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header:
            col_letter = get_column_letter(col)
            print(f"  {col_letter}: {header}")

    print("\n" + "-"*80)

    # Show all cells with content
    for row_idx in range(1, ws.max_row + 1):
        row_has_content = False
        row_data = []

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                row_has_content = True
                col_letter = get_column_letter(col_idx)

                # Get cell details
                cell_info = f"{col_letter}{row_idx}: {cell.value}"

                # Check for color
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                    rgb = cell.fill.fgColor.rgb
                    if rgb and rgb != "00000000":  # Not transparent
                        cell_info += f" [Color: #{rgb}]"

                # Check for comment
                if cell.comment:
                    cell_info += f" [Comment: {cell.comment.text}]"

                row_data.append(cell_info)

        if row_has_content:
            print(f"\nRow {row_idx}:")
            for data in row_data:
                print(f"  {data}")

    # Also search for specific patterns
    search_patterns = ["缩短", "Simon", "JIJIA"]
    print("\n" + "="*80)
    print("Searching for specific patterns:", search_patterns)

    for pattern in search_patterns:
        print(f"\nSearching for: '{pattern}'")
        print("-"*40)
        found = False

        for row in ws.iter_rows():
            for cell in row:
                if cell.value and pattern.lower() in str(cell.value).lower():
                    col_letter = get_column_letter(cell.column)
                    print(f"\nFound at: {col_letter}{cell.row}")
                    print(f"Cell value: {cell.value}")

                    # Show context (surrounding cells)
                    print("\nContext:")
                    # Show header (row 1)
                    header = ws.cell(row=1, column=cell.column).value
                    if header:
                        print(f"  Column header: {header}")

                    # Show row label (column A)
                    row_label = ws.cell(row=cell.row, column=1).value
                    if row_label:
                        print(f"  Row label: {row_label}")

                    # Check cell color
                    if cell.fill and cell.fill.fgColor.rgb:
                        print(f"  Cell color: #{cell.fill.fgColor.rgb}")

                    # Check cell comment
                    if cell.comment:
                        print(f"  Comment: {cell.comment.text}")

                    found = True

        if not found:
            print(f"Text '{search_text}' not found in PET sheet")

    else:
        # Show first 20 rows with content
        print("\nShowing content (max 20 rows with data):")
        print("="*80)

        rows_shown = 0
        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            if rows_shown >= 20:
                break

            row_data = []
            has_content = False

            for col_idx, cell in enumerate(row, start=1):
                if cell.value:
                    has_content = True
                    col_letter = get_column_letter(col_idx)
                    # Truncate long values
                    value_str = str(cell.value)
                    if len(value_str) > 50:
                        value_str = value_str[:47] + "..."
                    row_data.append(f"{col_letter}: {value_str}")

            if has_content:
                print(f"\nRow {row_idx}:")
                for data in row_data:
                    print(f"  {data}")
                rows_shown += 1

        print(f"\n(Showing {rows_shown} rows with content)")

    wb.close()

if __name__ == "__main__":
    check_pet_sheet()
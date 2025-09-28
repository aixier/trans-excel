"""Excel loader service."""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from typing import Optional, Dict, Any
import uuid
from pathlib import Path

from models.excel_dataframe import ExcelDataFrame


class ExcelLoader:
    """Load Excel files with color and comment extraction."""

    @staticmethod
    def load_excel(file_path: str) -> ExcelDataFrame:
        """Load Excel file with all metadata."""
        excel_df = ExcelDataFrame()
        excel_df.filename = Path(file_path).name
        excel_df.excel_id = str(uuid.uuid4())

        # Load workbook for metadata extraction
        wb = openpyxl.load_workbook(file_path, data_only=False)

        # Load all sheets
        excel_data = pd.read_excel(file_path, sheet_name=None)

        for sheet_name, df in excel_data.items():
            # Add DataFrame
            excel_df.add_sheet(sheet_name, df)

            # Get worksheet
            ws = wb[sheet_name]

            # Extract colors and comments
            for row_idx, row in enumerate(ws.iter_rows()):
                for col_idx, cell in enumerate(row):
                    # Extract color
                    if cell.fill and cell.fill.patternType:
                        fill = cell.fill
                        if isinstance(fill.fgColor.rgb, str):
                            color = f"#{fill.fgColor.rgb}" if fill.fgColor.rgb else None
                            if color and color != "#00000000":  # Ignore transparent
                                # Skip header row for DataFrame indexing (pandas read_excel skips header)
                                if row_idx > 0:
                                    excel_df.set_cell_color(sheet_name, row_idx - 1, col_idx, color)

                    # Extract comment
                    if cell.comment:
                        # Skip header row for DataFrame indexing
                        if row_idx > 0:
                            excel_df.set_cell_comment(
                                sheet_name, row_idx - 1, col_idx, cell.comment.text
                            )

        wb.close()
        return excel_df

    @staticmethod
    def save_excel(excel_df: ExcelDataFrame, file_path: str) -> None:
        """Save Excel file with colors and comments."""
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Write all sheets
            for sheet_name, df in excel_df.sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Get workbook and apply colors/comments
            wb = writer.book

            for sheet_name in excel_df.get_sheet_names():
                ws = wb[sheet_name]

                # Apply colors
                if sheet_name in excel_df.color_map:
                    for (row, col), color in excel_df.color_map[sheet_name].items():
                        cell = ws.cell(row=row+2, column=col+1)  # +2 for header, +1 for 1-based
                        if color and color.startswith('#'):
                            fill = PatternFill(
                                start_color=color[1:],
                                end_color=color[1:],
                                fill_type='solid'
                            )
                            cell.fill = fill

                # Apply comments
                if sheet_name in excel_df.comment_map:
                    for (row, col), comment_text in excel_df.comment_map[sheet_name].items():
                        cell = ws.cell(row=row+2, column=col+1)  # +2 for header, +1 for 1-based
                        cell.comment = openpyxl.comments.Comment(comment_text, "System")

    @staticmethod
    def validate_excel(file_path: str) -> Dict[str, Any]:
        """Validate Excel file format."""
        try:
            # Try to load the file
            excel_data = pd.read_excel(file_path, sheet_name=None, nrows=5)

            return {
                'valid': True,
                'sheets': list(excel_data.keys()),
                'sheet_count': len(excel_data)
            }
        except Exception as e:
            return {
                'valid': False,
                'error': str(e)
            }
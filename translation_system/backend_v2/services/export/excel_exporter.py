"""Excel export optimization service."""

import os
import logging
from pathlib import Path
from typing import Dict, Any, Optional, List
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from datetime import datetime

from models.excel_dataframe import ExcelDataFrame
from models.task_dataframe import TaskDataFrameManager, TaskStatus
from utils.session_manager import session_manager


logger = logging.getLogger(__name__)


class ExcelExporter:
    """Excel exporter with format preservation and optimization."""

    def __init__(self, output_dir: str = None):
        """
        Initialize ExcelExporter.

        Args:
            output_dir: Output directory for exported files
        """
        self.output_dir = Path(output_dir) if output_dir else Path("./output")
        self.output_dir.mkdir(exist_ok=True)
        self.logger = logging.getLogger(self.__class__.__name__)

    async def export_final_excel(self, session_id: str) -> str:
        """
        Export final Excel file with translation results.

        Args:
            session_id: Session identifier

        Returns:
            Path to the exported Excel file
        """
        try:
            # Get session data
            excel_df = session_manager.get_excel_df(session_id)
            task_manager = session_manager.get_task_manager(session_id)

            if not excel_df or not task_manager:
                raise ValueError(f"Session {session_id} not found or incomplete")

            # Create output filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_filename = Path(excel_df.filename).stem
            output_filename = f"{base_filename}_translated_{timestamp}.xlsx"
            output_path = self.output_dir / output_filename

            self.logger.info(f"Exporting Excel file to: {output_path}")

            # Create new workbook
            wb = Workbook()

            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # Process each sheet
            for sheet_name, df in excel_df.sheets.items():
                ws = wb.create_sheet(title=sheet_name)

                # Write data with translations
                await self._write_sheet_with_translations(
                    ws, df, excel_df, task_manager, sheet_name
                )

                # Apply formatting
                self._apply_sheet_formatting(ws, df, excel_df, sheet_name)

            # Save workbook
            wb.save(output_path)

            # Store export metadata
            session_manager.set_metadata(
                session_id, 'exported_file', str(output_path)
            )
            session_manager.set_metadata(
                session_id, 'export_timestamp', timestamp
            )

            self.logger.info(f"Successfully exported Excel file: {output_path}")
            return str(output_path)

        except Exception as e:
            self.logger.error(f"Failed to export Excel for session {session_id}: {e}")
            raise

    async def _write_sheet_with_translations(
        self,
        worksheet,
        df: pd.DataFrame,
        excel_df: ExcelDataFrame,
        task_manager: TaskDataFrameManager,
        sheet_name: str
    ):
        """Write sheet data with translation results."""

        # Get completed tasks for this sheet
        if task_manager.df is not None:
            sheet_tasks = task_manager.df[
                (task_manager.df['sheet_name'] == sheet_name) &
                (task_manager.df['status'] == TaskStatus.COMPLETED)
            ]
        else:
            sheet_tasks = pd.DataFrame()

        # Create translation map: (row, col) -> translated_text
        translation_map = {}
        for _, task in sheet_tasks.iterrows():
            row_idx = int(task['row_idx'])
            col_idx = int(task['col_idx'])
            translation_map[(row_idx, col_idx)] = task['result']

        # Write data row by row
        for row_idx in range(len(df)):
            for col_idx in range(len(df.columns)):
                # Get original value
                original_value = df.iloc[row_idx, col_idx]

                # Use translation if available, otherwise use original
                if (row_idx, col_idx) in translation_map:
                    cell_value = translation_map[(row_idx, col_idx)]
                else:
                    cell_value = original_value

                # Write to Excel (1-based indexing)
                excel_row = row_idx + 1
                excel_col = col_idx + 1
                cell = worksheet.cell(row=excel_row, column=excel_col)
                cell.value = cell_value

                # Add comment if task was translated
                if (row_idx, col_idx) in translation_map:
                    original_comment = excel_df.get_cell_comment(
                        sheet_name, row_idx, col_idx
                    )
                    translation_comment = f"Translated from: {original_value}"

                    if original_comment:
                        final_comment = f"{original_comment}\n{translation_comment}"
                    else:
                        final_comment = translation_comment

                    cell.comment = Comment(final_comment, "TranslationSystem")

    def _apply_sheet_formatting(
        self,
        worksheet,
        df: pd.DataFrame,
        excel_df: ExcelDataFrame,
        sheet_name: str
    ):
        """Apply formatting to worksheet."""

        # Create named styles
        translated_style = NamedStyle(name="translated")
        translated_style.fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
        translated_style.font = Font(italic=True)

        # Apply cell colors and formatting
        for row_idx in range(len(df)):
            for col_idx in range(len(df.columns)):
                excel_row = row_idx + 1
                excel_col = col_idx + 1
                cell = worksheet.cell(row=excel_row, column=excel_col)

                # Apply original color if exists
                original_color = excel_df.get_cell_color(
                    sheet_name, row_idx, col_idx
                )

                if original_color:
                    # Remove '#' if present
                    color = original_color.lstrip('#')
                    fill = PatternFill(
                        start_color=color, end_color=color, fill_type="solid"
                    )
                    cell.fill = fill

                # Mark translated cells
                if cell.comment:
                    # Add gray background to translated cells
                    current_fill = cell.fill
                    if not current_fill or current_fill.start_color.rgb == '00000000':
                        cell.fill = translated_style.fill
                    cell.font = Font(italic=True)

        # Auto-adjust column widths
        self._auto_adjust_columns(worksheet, df)

    def _auto_adjust_columns(self, worksheet, df: pd.DataFrame):
        """Auto-adjust column widths based on content."""

        for col_idx in range(len(df.columns)):
            column_letter = get_column_letter(col_idx + 1)

            # Calculate max width for column
            max_width = 10  # Minimum width

            for row_idx in range(len(df)):
                cell_value = df.iloc[row_idx, col_idx]
                if pd.notna(cell_value):
                    cell_length = len(str(cell_value))
                    max_width = max(max_width, cell_length)

            # Set width (with reasonable limits)
            width = min(max_width + 2, 50)  # Max 50 characters
            worksheet.column_dimensions[column_letter].width = width

    async def export_task_summary(self, session_id: str) -> str:
        """
        Export task summary Excel file.

        Args:
            session_id: Session identifier

        Returns:
            Path to the summary Excel file
        """
        try:
            task_manager = session_manager.get_task_manager(session_id)

            if not task_manager or task_manager.df is None:
                raise ValueError(f"No tasks found for session {session_id}")

            # Create output filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            summary_filename = f"task_summary_{session_id[:8]}_{timestamp}.xlsx"
            summary_path = self.output_dir / summary_filename

            # Export task DataFrame to Excel
            with pd.ExcelWriter(summary_path, engine='openpyxl') as writer:
                # Main tasks sheet
                task_manager.df.to_excel(writer, sheet_name='Tasks', index=False)

                # Statistics sheet
                stats = task_manager.get_statistics()
                stats_df = pd.DataFrame([stats])
                stats_df.to_excel(writer, sheet_name='Statistics', index=False)

                # By status sheet
                if stats['by_status']:
                    status_df = pd.DataFrame(
                        list(stats['by_status'].items()),
                        columns=['Status', 'Count']
                    )
                    status_df.to_excel(writer, sheet_name='By_Status', index=False)

            self.logger.info(f"Exported task summary to: {summary_path}")
            return str(summary_path)

        except Exception as e:
            self.logger.error(f"Failed to export task summary: {e}")
            raise

    def get_export_info(self, session_id: str) -> Dict[str, Any]:
        """
        Get export information for a session.

        Args:
            session_id: Session identifier

        Returns:
            Export information
        """
        exported_file = session_manager.get_metadata(session_id, 'exported_file')
        export_timestamp = session_manager.get_metadata(session_id, 'export_timestamp')

        info = {
            'has_export': exported_file is not None,
            'exported_file': exported_file,
            'export_timestamp': export_timestamp,
            'file_exists': False,
            'file_size': None
        }

        if exported_file and os.path.exists(exported_file):
            info['file_exists'] = True
            info['file_size'] = os.path.getsize(exported_file)

        return info

    async def cleanup_old_exports(self, days_old: int = 7):
        """
        Clean up old exported files.

        Args:
            days_old: Remove files older than this many days
        """
        try:
            if not self.output_dir.exists():
                return

            current_time = datetime.now()
            cleanup_count = 0

            for file_path in self.output_dir.glob("*.xlsx"):
                # Check file age
                file_time = datetime.fromtimestamp(file_path.stat().st_mtime)
                age_days = (current_time - file_time).days

                if age_days > days_old:
                    try:
                        file_path.unlink()
                        cleanup_count += 1
                        self.logger.debug(f"Removed old export file: {file_path}")
                    except Exception as e:
                        self.logger.warning(f"Failed to remove file {file_path}: {e}")

            if cleanup_count > 0:
                self.logger.info(f"Cleaned up {cleanup_count} old export files")

        except Exception as e:
            self.logger.error(f"Failed to cleanup old exports: {e}")


# Global exporter instance
excel_exporter = ExcelExporter()
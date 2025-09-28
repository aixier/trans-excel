#!/usr/bin/env python3
"""Debug script to understand task counting logic"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from utils.color_detector import is_yellow_color, is_blue_color

def analyze_excel():
    file_path = '/mnt/d/work/trans_excel/sfdaf.xlsx'

    # Load workbook for colors
    workbook = load_workbook(file_path, data_only=True)

    # Load with pandas for data
    sheet_name = workbook.sheetnames[0]
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=0)

    print(f"Sheet: {sheet_name}")
    print(f"Shape: {df.shape}")
    print(f"Columns: {list(df.columns)}")
    print()

    # Get the actual sheet
    sheet = workbook[sheet_name]

    # Count yellow cells and track unique source texts
    yellow_count = 0
    blue_count = 0
    yellow_texts = set()
    blue_texts = set()
    empty_tr_count = 0

    # Find column indices
    columns = list(df.columns)
    en_col_idx = None
    tr_col_idx = None

    for idx, col in enumerate(columns):
        col_upper = str(col).upper()
        if col_upper == 'EN':
            en_col_idx = idx
        elif col_upper == 'TR':
            tr_col_idx = idx

    print(f"EN column index: {en_col_idx}")
    print(f"TR column index: {tr_col_idx}")
    print()

    # Iterate through rows
    for row_idx in range(len(df)):
        # Check EN column (source)
        if en_col_idx is not None:
            en_value = df.iloc[row_idx, en_col_idx]
            if pd.notna(en_value) and str(en_value).strip():
                # Get the cell color (row+2 because of header and 1-based indexing)
                cell = sheet.cell(row=row_idx+2, column=en_col_idx+1)

                if cell.fill and hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
                    color_value = cell.fill.fgColor.rgb
                    if color_value and isinstance(color_value, str):
                        color_hex = f"#{color_value[2:]}" if len(color_value) == 8 else f"#{color_value}"

                        if is_yellow_color(color_hex):
                            yellow_count += 1
                            yellow_texts.add(str(en_value).strip())
                            if row_idx < 5:  # Print first few
                                print(f"Row {row_idx+2}: Yellow EN cell: '{en_value[:50]}...' Color: {color_hex}")
                        elif is_blue_color(color_hex):
                            blue_count += 1
                            blue_texts.add(str(en_value).strip())

        # Check TR column (target) for empty cells
        if tr_col_idx is not None and en_col_idx is not None:
            tr_value = df.iloc[row_idx, tr_col_idx]
            en_value = df.iloc[row_idx, en_col_idx]

            # If TR is empty but EN has content
            if (pd.isna(tr_value) or str(tr_value).strip() == '') and pd.notna(en_value) and str(en_value).strip():
                empty_tr_count += 1

    print(f"\n=== Results ===")
    print(f"Total rows with data: {len(df)}")
    print(f"Yellow EN cells: {yellow_count}")
    print(f"Unique yellow texts: {len(yellow_texts)}")
    print(f"Blue EN cells: {blue_count}")
    print(f"Unique blue texts: {len(blue_texts)}")
    print(f"Empty TR cells (with EN content): {empty_tr_count}")
    print(f"\nExpected tasks (yellow cells): 84")
    print(f"Actual yellow cells found: {yellow_count}")

    # Analyze why the count might be different
    if yellow_count != 84:
        print(f"\n=== Debugging ===")
        # Check a specific row we know should be yellow
        test_row = 1  # Row 2 in Excel (0-based in pandas)
        if en_col_idx is not None:
            cell = sheet.cell(row=test_row+2, column=en_col_idx+1)
            if cell.fill and hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
                color_value = cell.fill.fgColor.rgb
                if color_value:
                    color_hex = f"#{color_value[2:]}" if len(color_value) == 8 else f"#{color_value}"
                    print(f"Sample cell (row {test_row+2}): Color={color_hex}")
                    print(f"Is yellow? {is_yellow_color(color_hex)}")

                    # Parse RGB values
                    hex_color = color_hex.lstrip('#')
                    r = int(hex_color[0:2], 16)
                    g = int(hex_color[2:4], 16)
                    b = int(hex_color[4:6], 16)
                    print(f"RGB values: R={r}, G={g}, B={b}")
                    print(f"Yellow criteria: R>180:{r>180}, G>180:{g>180}, B<150:{b<150}")

if __name__ == '__main__':
    analyze_excel()
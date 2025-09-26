"""Quick test focusing on yellow cells."""

import openpyxl
import pandas as pd
from datetime import datetime

def quick_test():
    """Quick test focusing on yellow cells that need translation."""

    file_path = "/mnt/d/work/trans_excel/test_text_targ3_5tab_normall_有注释.xlsx"

    print(f"\n{'='*60}")
    print("Quick Analysis - Yellow Cells (Translation Tasks)")
    print('='*60)

    # Load workbook
    print("\nLoading workbook...")
    start_time = datetime.now()
    wb = openpyxl.load_workbook(file_path, data_only=True)  # data_only for faster loading
    print(f"Loaded in {(datetime.now() - start_time).total_seconds():.2f} seconds")

    total_yellow = 0
    tasks = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_yellow = 0

        # Only check first 100 rows for quick test
        max_row = min(ws.max_row, 100)

        for row in range(1, max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)

                # Check if cell is yellow
                if cell.fill and cell.fill.patternType:
                    if hasattr(cell.fill.fgColor, 'rgb') and cell.fill.fgColor.rgb:
                        color = str(cell.fill.fgColor.rgb)
                        if 'FFFF' in color or 'ffff' in color:
                            # Found yellow cell
                            sheet_yellow += 1

                            # Get source text (assume column B is CH, C is EN)
                            source_ch = ws.cell(row=row, column=2).value
                            source_en = ws.cell(row=row, column=3).value

                            tasks.append({
                                'sheet': sheet_name,
                                'row': row,
                                'col': col,
                                'cell_ref': f"{openpyxl.utils.get_column_letter(col)}{row}",
                                'value': cell.value,
                                'source_ch': source_ch,
                                'source_en': source_en,
                                'comment': cell.comment.text if cell.comment else None
                            })

        if sheet_yellow > 0:
            print(f"\n[{sheet_name}] Found {sheet_yellow} yellow cells")
        total_yellow += sheet_yellow

    wb.close()

    print(f"\n{'='*60}")
    print(f"Summary:")
    print(f"  Total yellow cells: {total_yellow}")
    print(f"  Translation tasks needed: {total_yellow * 3} (for TH, PT, VN)")

    # Show first 5 tasks
    if tasks:
        print(f"\nFirst 5 translation tasks:")
        for i, task in enumerate(tasks[:5]):
            print(f"\n  Task {i+1}:")
            print(f"    Location: {task['sheet']}!{task['cell_ref']}")
            print(f"    EN Source: {task['source_en']}")
            if task['comment']:
                print(f"    Comment: {task['comment']}")

    # Calculate batches
    avg_chars = 100  # Assume average 100 chars per task
    max_chars_per_batch = 50000
    estimated_batches = (total_yellow * 3 * avg_chars) // max_chars_per_batch + 1

    print(f"\n  Estimated batches: {estimated_batches}")
    print(f"  (Based on {max_chars_per_batch} chars per batch)")

    print(f"\n{'='*60}")
    print("Quick analysis complete!")

if __name__ == "__main__":
    quick_test()
"""Create test Excel files for testing."""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
import os


def create_small_test_file():
    """Create a small test Excel file with 10 rows."""
    data = {
        'ID': range(1, 11),
        'CH': ['开始游戏', '设置', '退出', '背包', '商店',
               '任务', '地图', '角色', '技能', '装备'],
        'EN': ['Start Game', 'Settings', 'Exit', 'Inventory', 'Shop',
               'Quest', 'Map', 'Character', 'Skills', 'Equipment'],
        'PT': ['', '', '', '', '', '', '', '', '', ''],  # Empty for translation
        'TH': ['', '', '', '', '', '', '', '', '', ''],  # Empty for translation
        'VN': ['', '', '', '', '', '', '', '', '', '']   # Empty for translation
    }

    df = pd.DataFrame(data)

    # Save with colors
    file_path = 'small.xlsx'
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='UI_Text', index=False)

        # Get worksheet
        ws = writer.sheets['UI_Text']

        # Add yellow color to some PT cells (need translation)
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        for row in range(2, 7):  # Rows 2-6 (Excel 1-indexed, plus header)
            ws.cell(row=row, column=4).fill = yellow_fill  # PT column

        # Add blue color to some TH cells
        blue_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
        for row in range(2, 5):  # Rows 2-4
            ws.cell(row=row, column=5).fill = blue_fill  # TH column

        # Add comments
        ws.cell(row=2, column=2).comment = Comment('Main menu button', 'System')
        ws.cell(row=3, column=2).comment = Comment('Game settings', 'System')

    print(f"Created {file_path}")
    return file_path


def create_medium_test_file():
    """Create a medium test Excel file with 100 rows."""
    import random

    categories = ['UI', 'Dialog', 'Item', 'Skill', 'Quest']
    ch_texts = [
        '确认', '取消', '保存', '加载', '删除',
        '你好，冒险者！', '欢迎来到我们的村庄。', '有什么需要帮助的吗？',
        '铁剑', '皮甲', '生命药水', '魔法卷轴',
        '火球术', '治疗术', '闪电链', '冰霜新星',
        '击败10个哥布林', '收集5个草药', '护送商人到下一个城镇'
    ]

    data = {
        'Category': [],
        'ID': [],
        'CH': [],
        'EN': [],
        'PT': [],
        'TH': [],
        'VN': []
    }

    for i in range(100):
        category = random.choice(categories)
        data['Category'].append(category)
        data['ID'].append(f"{category}_{i:03d}")
        data['CH'].append(random.choice(ch_texts))
        data['EN'].append(f"English text {i}")
        data['PT'].append('')
        data['TH'].append('')
        data['VN'].append('')

    df = pd.DataFrame(data)

    # Create multiple sheets
    file_path = 'medium.xlsx'
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        # Split by category
        for category in categories:
            cat_df = df[df['Category'] == category].reset_index(drop=True)
            cat_df.to_excel(writer, sheet_name=category, index=False)

            # Add some colors
            ws = writer.sheets[category]
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            for row in range(2, min(len(cat_df) + 2, 10)):
                ws.cell(row=row, column=5).fill = yellow_fill  # PT column

    print(f"Created {file_path}")
    return file_path


def create_large_test_file():
    """Create a large test Excel file with 1000 rows."""
    data = {
        'ID': range(1, 1001),
        'Source_CH': [f'中文文本{i}' for i in range(1, 1001)],
        'Source_EN': [f'English text {i}' for i in range(1, 1001)],
        'Target_PT': ['' for _ in range(1000)],
        'Target_TH': ['' for _ in range(1000)],
        'Target_VN': ['' for _ in range(1000)]
    }

    df = pd.DataFrame(data)
    file_path = 'large.xlsx'
    df.to_excel(file_path, sheet_name='Translations', index=False)
    print(f"Created {file_path}")
    return file_path


def create_mixed_test_file():
    """Create a test file with mixed character lengths."""
    data = {
        'ID': [],
        'Text_CH': [],
        'Text_EN': [],
        'PT': [],
        'TH': [],
        'VN': []
    }

    # Add texts of various lengths
    texts = [
        ('短', 'Short'),  # 1 char
        ('中等长度文本', 'Medium length text'),  # ~10 chars
        ('这是一个比较长的文本，用于测试较长内容的翻译。', 'This is a longer text for testing translation of longer content.'),  # ~30 chars
        ('这是一段非常长的文本内容。' * 10, 'This is a very long text content. ' * 10),  # ~130 chars
    ]

    for i, (ch, en) in enumerate(texts * 25):  # 100 rows total
        data['ID'].append(i + 1)
        data['Text_CH'].append(ch)
        data['Text_EN'].append(en)
        data['PT'].append('')
        data['TH'].append('')
        data['VN'].append('')

    df = pd.DataFrame(data)
    file_path = 'mixed.xlsx'

    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Mixed_Content', index=False)

        # Add various colors
        ws = writer.sheets['Mixed_Content']
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        blue_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')

        for row in range(2, len(df) + 2):
            if row % 3 == 0:
                ws.cell(row=row, column=4).fill = yellow_fill  # PT
            elif row % 3 == 1:
                ws.cell(row=row, column=5).fill = blue_fill  # TH

    print(f"Created {file_path}")
    return file_path


if __name__ == "__main__":
    # Create test data directory
    os.makedirs('tests/test_data', exist_ok=True)
    os.chdir('tests/test_data')

    # Create all test files
    create_small_test_file()
    create_medium_test_file()
    create_large_test_file()
    create_mixed_test_file()

    print("All test files created successfully!")
"""Test color and comment extraction from Excel."""

import openpyxl
import pandas as pd
import sys
import os

def analyze_excel_colors_comments(file_path):
    """Analyze colors and comments in Excel file using openpyxl."""
    print(f"\n{'='*60}")
    print(f"Analyzing Excel File with openpyxl")
    print(f"File: {os.path.basename(file_path)}")
    print('='*60)

    try:
        # Load workbook
        print("\nLoading workbook...")
        wb = openpyxl.load_workbook(file_path, data_only=False)
        print(f"Sheets: {wb.sheetnames}")

        for sheet_name in wb.sheetnames:
            print(f"\n[Sheet: {sheet_name}]")
            ws = wb[sheet_name]
            print(f"  Dimensions: {ws.max_row} rows x {ws.max_column} columns")

            # Analyze colors
            yellow_cells = []
            blue_cells = []
            other_colored_cells = []

            for row in ws.iter_rows():
                for cell in row:
                    if cell.fill and cell.fill.patternType:
                        if hasattr(cell.fill.fgColor, 'rgb') and cell.fill.fgColor.rgb:
                            color = cell.fill.fgColor.rgb

                            # Check for yellow (FFFF00 or similar)
                            if 'FFFF' in str(color) or 'ffff' in str(color):
                                yellow_cells.append((cell.row, cell.column, cell.value))
                            # Check for blue (0000FF or similar)
                            elif '0000FF' in str(color) or '0000ff' in str(color):
                                blue_cells.append((cell.row, cell.column, cell.value))
                            elif color != '00000000':  # Not transparent
                                other_colored_cells.append((cell.row, cell.column, color, cell.value))

            print(f"\n  Color Analysis:")
            print(f"    - Yellow cells: {len(yellow_cells)}")
            if yellow_cells:
                for row, col, value in yellow_cells[:5]:  # Show first 5
                    col_letter = openpyxl.utils.get_column_letter(col)
                    print(f"      {col_letter}{row}: {value}")

            print(f"    - Blue cells: {len(blue_cells)}")
            if blue_cells:
                for row, col, value in blue_cells[:5]:  # Show first 5
                    col_letter = openpyxl.utils.get_column_letter(col)
                    print(f"      {col_letter}{row}: {value}")

            if other_colored_cells:
                print(f"    - Other colored cells: {len(other_colored_cells)}")
                for row, col, color, value in other_colored_cells[:3]:
                    col_letter = openpyxl.utils.get_column_letter(col)
                    print(f"      {col_letter}{row}: #{color} = {value}")

            # Analyze comments
            cells_with_comments = []
            for row in ws.iter_rows():
                for cell in row:
                    if cell.comment:
                        cells_with_comments.append((cell.row, cell.column, cell.comment.text, cell.value))

            print(f"\n  Comment Analysis:")
            print(f"    - Cells with comments: {len(cells_with_comments)}")
            if cells_with_comments:
                for row, col, comment, value in cells_with_comments[:5]:  # Show first 5
                    col_letter = openpyxl.utils.get_column_letter(col)
                    print(f"      {col_letter}{row}: \"{comment}\" (value: {value})")

        wb.close()
        print("\n✓ Analysis complete")

    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()


def test_with_pandas(file_path):
    """Quick test with pandas to see data structure."""
    print(f"\n{'='*60}")
    print("Quick Data Preview with Pandas")
    print('='*60)

    try:
        excel_data = pd.read_excel(file_path, sheet_name=None)

        for sheet_name, df in excel_data.items():
            print(f"\n[Sheet: {sheet_name}]")
            print(f"  Shape: {df.shape}")
            print(f"  Columns: {list(df.columns)}")

            # Show first few non-null values from each column
            print("\n  Sample data:")
            for col in df.columns[:6]:  # First 6 columns
                non_null = df[col].dropna().head(3)
                if not non_null.empty:
                    print(f"    {col}: {non_null.tolist()}")

    except Exception as e:
        print(f"Error with pandas: {e}")


if __name__ == "__main__":
    test_file = "/mnt/d/work/trans_excel/test_text_targ3_5tab_normall_有注释.xlsx"

    if not os.path.exists(test_file):
        print(f"File not found: {test_file}")
        sys.exit(1)

    print(f"\n{'#'*60}")
    print("# Excel Color and Comment Analysis")
    print(f"{'#'*60}")

    # First get quick overview with pandas
    test_with_pandas(test_file)

    # Then analyze colors and comments with openpyxl
    analyze_excel_colors_comments(test_file)

    print(f"\n{'='*60}")
    print("Analysis completed!")
    print('='*60)
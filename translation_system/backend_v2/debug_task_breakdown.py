#!/usr/bin/env python3
"""Debug script to understand task breakdown logic"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from utils.color_detector import is_yellow_color, is_blue_color

def analyze_tasks():
    file_path = '/mnt/d/work/trans_excel/sfdaf.xlsx'

    # Load workbook for colors
    workbook = load_workbook(file_path, data_only=True)

    # Load with pandas for data
    sheet_name = workbook.sheetnames[0]
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=0)

    print(f"Sheet: {sheet_name}")
    print(f"Shape: {df.shape}")
    print(f"Columns: {list(df.columns)}")
    print()

    # Get the actual sheet
    sheet = workbook[sheet_name]

    # Identify source and target columns
    columns = list(df.columns)
    source_cols = []
    target_cols = []

    for idx, col in enumerate(columns):
        col_upper = str(col).upper()
        # Source columns
        if col_upper in ['CH', 'CN', '中文', 'EN', 'ENGLISH', '英文']:
            source_cols.append(idx)
            print(f"Source column: {col} (index {idx})")
        # Target columns
        elif col_upper in ['TR', 'TH', 'PT', 'VN', 'VI', 'ES', 'IND', 'ID', 'TURKISH', '土耳其语']:
            target_cols.append(idx)
            print(f"Target column: {col} (index {idx})")

    print()

    # Count tasks
    normal_tasks = 0
    yellow_tasks = 0
    blue_tasks = 0
    yellow_cells = []
    blue_cells = []
    normal_cells = []

    for row_idx in range(len(df)):
        for col_idx in range(len(columns)):
            cell_value = df.iloc[row_idx, col_idx]

            if pd.notna(cell_value) and isinstance(cell_value, str) and len(cell_value.strip()) > 0:
                # Get cell color (row+2 because of header and 1-based indexing)
                cell = sheet.cell(row=row_idx+2, column=col_idx+1)
                cell_color = None

                if cell.fill and hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
                    color_value = cell.fill.fgColor.rgb
                    if color_value and isinstance(color_value, str):
                        cell_color = f"#{color_value[2:]}" if len(color_value) == 8 else f"#{color_value}"

                text = str(cell_value).strip()

                # Type 3: Blue cells (shortening to self)
                if cell_color and is_blue_color(cell_color):
                    blue_tasks += 1
                    blue_cells.append(f"Row {row_idx+2}, Col {columns[col_idx]}: {text[:30]}")

                # Type 2: Yellow cells (re-translate to all following columns)
                elif cell_color and is_yellow_color(cell_color):
                    cols_after = len(columns) - col_idx - 1
                    if cols_after > 0:
                        yellow_tasks += cols_after
                        yellow_cells.append(f"Row {row_idx+2}, Col {columns[col_idx]}: {text[:30]} -> {cols_after} columns after")

                # Type 1: Normal translation (source to targets)
                elif col_idx in source_cols:
                    # Check if any target columns are empty
                    for target_col_idx in target_cols:
                        if target_col_idx < len(columns):
                            target_value = df.iloc[row_idx, target_col_idx]
                            if pd.isna(target_value) or str(target_value).strip() == '':
                                normal_tasks += 1
                                if len(normal_cells) < 5:  # Only show first 5
                                    normal_cells.append(f"Row {row_idx+2}: {columns[col_idx]} -> {columns[target_col_idx]} (empty)")

    print("=== Task Breakdown ===")
    print(f"Normal tasks: {normal_tasks}")
    print(f"Yellow tasks: {yellow_tasks}")
    print(f"Blue tasks: {blue_tasks}")
    print(f"Total: {normal_tasks + yellow_tasks + blue_tasks}")
    print()

    print("=== Sample Normal Tasks ===")
    for cell in normal_cells[:5]:
        print(f"  {cell}")

    print()
    print("=== Sample Yellow Tasks ===")
    for cell in yellow_cells[:5]:
        print(f"  {cell}")

    print()
    print("=== Sample Blue Tasks ===")
    for cell in blue_cells[:5]:
        print(f"  {cell}")

    # Verify the numbers
    print()
    print("=== Verification ===")
    print(f"Total rows with data: {len(df)}")
    print(f"Source columns: {[columns[i] for i in source_cols]}")
    print(f"Target columns: {[columns[i] for i in target_cols]}")

    # Count non-empty EN cells
    if 2 in source_cols:  # EN column
        en_non_empty = 0
        for row_idx in range(len(df)):
            en_value = df.iloc[row_idx, 2]  # EN column
            if pd.notna(en_value) and str(en_value).strip():
                en_non_empty += 1
        print(f"Non-empty EN cells: {en_non_empty}")

    # Count empty TR cells
    if 3 in target_cols:  # TR column
        tr_empty = 0
        for row_idx in range(len(df)):
            tr_value = df.iloc[row_idx, 3]  # TR column
            if pd.isna(tr_value) or str(tr_value).strip() == '':
                tr_empty += 1
        print(f"Empty TR cells: {tr_empty}")

if __name__ == '__main__':
    analyze_tasks()
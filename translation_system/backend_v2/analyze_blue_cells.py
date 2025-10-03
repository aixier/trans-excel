#!/usr/bin/env python3
"""Analyze blue cells in Excel files."""

import openpyxl
import pandas as pd

def analyze_blue_cells(file_path):
    """Analyze blue cells in an Excel file."""
    wb = openpyxl.load_workbook(file_path)
    blue_cells = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_blues = []

        # Get column headers
        headers = []
        for col in range(1, min(21, ws.max_column + 1)):
            headers.append(ws.cell(row=1, column=col).value)

        # Check cells
        for row in range(1, min(51, ws.max_row + 1)):
            for col in range(1, min(21, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)

                # Check if cell has fill color
                if cell.fill and cell.fill.start_color:
                    color = cell.fill.start_color.rgb
                    if color and color not in ['00000000', 'FFFFFFFF', None]:
                        # Extract RGB values
                        if len(color) == 8:  # ARGB format
                            r = int(color[2:4], 16)
                            g = int(color[4:6], 16)
                            b = int(color[6:8], 16)
                        else:  # RGB format
                            r = int(color[0:2], 16) if len(color) >= 2 else 0
                            g = int(color[2:4], 16) if len(color) >= 4 else 0
                            b = int(color[4:6], 16) if len(color) >= 6 else 0

                        # Check if it's blue-ish
                        if b > 150 and b > r and b > g:
                            cell_ref = f'{openpyxl.utils.get_column_letter(col)}{row}'
                            value = str(cell.value) if cell.value else ''
                            sheet_blues.append({
                                'cell': cell_ref,
                                'row': row,
                                'col': col,
                                'col_name': headers[col-1] if col-1 < len(headers) else f'Col{col}',
                                'value': value[:100],
                                'color': color,
                                'rgb': (r, g, b)
                            })

        if sheet_blues:
            blue_cells[sheet_name] = sheet_blues

    return blue_cells

def analyze_task_file(file_path):
    """Analyze the task splitting result file."""
    df = pd.read_excel(file_path)

    # Filter for blue tasks
    blue_tasks = df[df['task_type'] == 'blue'] if 'task_type' in df.columns else pd.DataFrame()

    return df, blue_tasks

# Analyze source file
print("=" * 80)
print("ANALYZING SOURCE EXCEL FILE")
print("=" * 80)

source_blues = analyze_blue_cells('/mnt/d/work/trans_excel/test1.xlsx')

for sheet_name, cells in source_blues.items():
    print(f"\nSheet: {sheet_name}")
    print(f"Found {len(cells)} blue cells:")

    # Group by column
    by_column = {}
    for cell_info in cells:
        col_name = cell_info['col_name']
        if col_name not in by_column:
            by_column[col_name] = []
        by_column[col_name].append(cell_info)

    for col_name, col_cells in by_column.items():
        print(f"\n  Column '{col_name}':")
        for cell_info in col_cells[:3]:  # Show first 3 per column
            print(f"    {cell_info['cell']}: '{cell_info['value']}'")
            print(f"      RGB: {cell_info['rgb']}")
        if len(col_cells) > 3:
            print(f"    ... and {len(col_cells) - 3} more blue cells in this column")

# Analyze task file
print("\n" + "=" * 80)
print("ANALYZING TASK SPLITTING RESULT")
print("=" * 80)

df, blue_tasks = analyze_task_file('/mnt/d/work/trans_excel/tasks_00056c4c-8ecc-438b-b752-02e012f75431.xlsx')

print(f"\nTotal tasks: {len(df)}")
print(f"Task type distribution:")
if 'task_type' in df.columns:
    print(df['task_type'].value_counts())
else:
    print("  No 'task_type' column found")

if not blue_tasks.empty:
    print(f"\nBlue tasks: {len(blue_tasks)}")
    print("\nSample blue tasks:")

    # Show first few blue tasks
    for idx, task in blue_tasks.head(5).iterrows():
        print(f"\n  Task ID: {task.get('task_id', 'N/A')}")
        print(f"  Sheet: {task.get('sheet_name', 'N/A')}")
        print(f"  Cell: {task.get('cell_ref', 'N/A')}")
        print(f"  Source Lang: {task.get('source_lang', 'N/A')} -> Target Lang: {task.get('target_lang', 'N/A')}")
        print(f"  Source Text: {str(task.get('source_text', ''))[:100]}")
else:
    print("\nNo blue tasks found in the result file!")

# Compare source blue cells with generated tasks
print("\n" + "=" * 80)
print("COMPARISON ANALYSIS")
print("=" * 80)

if 'sheet_name' in df.columns and 'cell_ref' in df.columns:
    # Check if blue source cells resulted in blue tasks
    for sheet_name, cells in source_blues.items():
        sheet_tasks = df[df['sheet_name'] == sheet_name] if not df.empty else pd.DataFrame()

        if not sheet_tasks.empty:
            print(f"\nSheet: {sheet_name}")
            print(f"  Blue cells in source: {len(cells)}")
            sheet_blue_tasks = sheet_tasks[sheet_tasks['task_type'] == 'blue'] if 'task_type' in sheet_tasks.columns else pd.DataFrame()
            print(f"  Blue tasks generated: {len(sheet_blue_tasks)}")

            # Check specific cells
            for cell_info in cells[:3]:  # Check first 3
                cell_ref = cell_info['cell']
                # Look for tasks referencing this cell
                cell_tasks = sheet_tasks[sheet_tasks['cell_ref'] == cell_ref] if 'cell_ref' in sheet_tasks.columns else pd.DataFrame()
                if not cell_tasks.empty:
                    print(f"    Cell {cell_ref}: {len(cell_tasks)} task(s), types: {cell_tasks['task_type'].tolist() if 'task_type' in cell_tasks.columns else 'N/A'}")
                else:
                    print(f"    Cell {cell_ref}: No tasks generated")
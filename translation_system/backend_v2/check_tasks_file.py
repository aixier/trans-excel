#!/usr/bin/env python3
"""Check tasks file to find duplicate Simon comments"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

def check_tasks_file():
    # Check the tasks file
    file_path = '/mnt/d/work/trans_excel/tasks_82086920-ec45-439f-8d25-07324b9f3c98.xlsx'

    print(f"Checking file: {file_path}")
    print("="*80)

    # Read with pandas first to see the structure
    df = pd.read_excel(file_path)
    print(f"\nDataFrame shape: {df.shape}")
    print(f"Columns: {list(df.columns)}")

    # Check if 'source_context' column exists
    if 'source_context' in df.columns:
        # Find all rows with Simon comment
        simon_rows = df[df['source_context'].str.contains('Simon', na=False)]
        print(f"\nFound {len(simon_rows)} rows containing 'Simon' in source_context")

        if len(simon_rows) > 0:
            print("\nDetails of rows with Simon comments:")
            print("-"*80)
            for idx, row in simon_rows.iterrows():
                print(f"\nRow {idx + 2}:")  # +2 for Excel row (header + 0-index)
                print(f"  task_id: {row.get('task_id', 'N/A')}")
                print(f"  source_text: {row.get('source_text', 'N/A')[:50]}...")
                print(f"  target_lang: {row.get('target_lang', 'N/A')}")
                print(f"  sheet_name: {row.get('sheet_name', 'N/A')}")
                print(f"  row_idx: {row.get('row_idx', 'N/A')}")
                print(f"  col_idx: {row.get('col_idx', 'N/A')}")
                print(f"  cell_ref: {row.get('cell_ref', 'N/A')}")
                print(f"  source_context: {row.get('source_context', 'N/A')}")
                print(f"  task_type: {row.get('task_type', 'N/A')}")

    # Also check with openpyxl to see the raw cell values
    print("\n" + "="*80)
    print("Checking with openpyxl for more details...")

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Find source_context column
    header_row = 1
    source_context_col = None

    for col in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=col).value == 'source_context':
            source_context_col = col
            break

    if source_context_col:
        print(f"source_context is in column {get_column_letter(source_context_col)}")

        # Count occurrences of the exact Simon comment
        simon_comment = "Simon:\n缩短这句话的翻译"
        simon_comment_alt = "Simon: 缩短这句话的翻译"

        count_exact = 0
        count_alt = 0
        rows_with_simon = []

        for row in range(2, ws.max_row + 1):  # Skip header
            cell_value = ws.cell(row=row, column=source_context_col).value
            if cell_value:
                if simon_comment in str(cell_value):
                    count_exact += 1
                    rows_with_simon.append(row)
                elif simon_comment_alt in str(cell_value):
                    count_alt += 1
                    rows_with_simon.append(row)

        print(f"\nFound '{simon_comment}' (with newline): {count_exact} times")
        print(f"Found '{simon_comment_alt}' (without newline): {count_alt} times")
        print(f"Total rows with Simon comment: {len(rows_with_simon)}")

        if rows_with_simon:
            print(f"Rows with Simon comment: {rows_with_simon[:10]}...")  # Show first 10

    wb.close()

if __name__ == "__main__":
    check_tasks_file()
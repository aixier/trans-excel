"""Excel loader service - loads Excel from URL or bytes."""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from typing import Optional, Dict, Any, Union
from io import BytesIO
import uuid
import logging

from models.excel_dataframe import ExcelDataFrame
from utils.http_client import http_client

logger = logging.getLogger(__name__)


class ExcelLoader:
    """Load Excel files from URL or bytes with color and comment extraction."""

    @staticmethod
    def load_from_url(url: str) -> ExcelDataFrame:
        """
        Load Excel file from HTTP URL.

        Args:
            url: HTTP URL to download Excel file from

        Returns:
            ExcelDataFrame with all data and metadata

        Raises:
            Exception: If download or loading fails
        """
        logger.info(f"Loading Excel from URL: {url}")

        # Download file
        file_data = http_client.download_file(url)

        # Load from BytesIO
        return ExcelLoader.load_from_bytes(file_data)

    @staticmethod
    def load_from_bytes(file_data: Union[BytesIO, bytes], filename: str = "uploaded.xlsx") -> ExcelDataFrame:
        """
        Load Excel file from bytes or BytesIO.

        Args:
            file_data: File data as bytes or BytesIO
            filename: Original filename (optional)

        Returns:
            ExcelDataFrame with all data and metadata

        Raises:
            Exception: If loading fails
        """
        logger.info(f"Loading Excel from bytes: {filename}")

        # Convert bytes to BytesIO if needed
        if isinstance(file_data, bytes):
            file_data = BytesIO(file_data)

        excel_df = ExcelDataFrame()
        excel_df.filename = filename
        excel_df.excel_id = str(uuid.uuid4())

        try:
            # Load workbook for metadata extraction
            file_data.seek(0)
            wb = openpyxl.load_workbook(file_data, data_only=False)

            # Load all sheets with pandas
            file_data.seek(0)
            excel_data = pd.read_excel(file_data, sheet_name=None)

            for sheet_name, df in excel_data.items():
                # Add DataFrame
                excel_df.add_sheet(sheet_name, df)

                # Get worksheet
                ws = wb[sheet_name]

                # Extract colors and comments
                for row_idx, row in enumerate(ws.iter_rows()):
                    for col_idx, cell in enumerate(row):
                        # Extract color
                        if cell.fill and cell.fill.patternType:
                            fill = cell.fill
                            if hasattr(fill.fgColor, 'rgb') and isinstance(fill.fgColor.rgb, str):
                                color = f"#{fill.fgColor.rgb}" if fill.fgColor.rgb else None
                                if color and color != "#00000000":  # Ignore transparent
                                    # Skip header row for DataFrame indexing (pandas read_excel skips header)
                                    if row_idx > 0:
                                        excel_df.set_cell_color(sheet_name, row_idx - 1, col_idx, color)

                        # Extract comment
                        if cell.comment:
                            # Skip header row for DataFrame indexing
                            if row_idx > 0:
                                excel_df.set_cell_comment(
                                    sheet_name, row_idx - 1, col_idx, cell.comment.text
                                )

            wb.close()
            logger.info(f"Successfully loaded Excel with {len(excel_data)} sheets")
            return excel_df

        except Exception as e:
            logger.error(f"Failed to load Excel: {e}")
            raise Exception(f"Failed to load Excel: {str(e)}")

    @staticmethod
    def validate_excel(file_data: Union[BytesIO, bytes]) -> Dict[str, Any]:
        """
        Validate Excel file format.

        Args:
            file_data: File data as bytes or BytesIO

        Returns:
            Validation result dictionary
        """
        try:
            # Convert bytes to BytesIO if needed
            if isinstance(file_data, bytes):
                file_data = BytesIO(file_data)

            # Try to load the file
            file_data.seek(0)
            excel_data = pd.read_excel(file_data, sheet_name=None, nrows=5)

            return {
                'valid': True,
                'sheets': list(excel_data.keys()),
                'sheet_count': len(excel_data)
            }
        except Exception as e:
            return {
                'valid': False,
                'error': str(e)
            }


# Global Excel loader instance
excel_loader = ExcelLoader()

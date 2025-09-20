#!/usr/bin/env python3
"""
Excel文件分析脚本 - 分析可能导致翻译卡住的数据内容
"""
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook
import re
import json
from datetime import datetime
import sys
import os

def analyze_excel_file(file_path):
    """分析Excel文件的详细信息"""
    print(f"\n{'='*60}")
    print(f"分析文件: {file_path}")
    print(f"{'='*60}")

    if not os.path.exists(file_path):
        print(f"❌ 文件不存在: {file_path}")
        return None

    try:
        # 使用openpyxl加载工作簿
        wb = load_workbook(file_path, data_only=True)
        sheet_analysis = {}

        for sheet_name in wb.sheetnames:
            print(f"\n📊 分析Sheet: {sheet_name}")
            ws = wb[sheet_name]

            # 获取工作表的实际使用范围
            max_row = ws.max_row
            max_col = ws.max_column

            print(f"   尺寸: {max_row} 行 x {max_col} 列")

            # 分析每个单元格
            cell_analysis = {
                'total_cells': 0,
                'non_empty_cells': 0,
                'text_cells': 0,
                'numeric_cells': 0,
                'formula_cells': 0,
                'long_text_cells': [],
                'special_char_cells': [],
                'problematic_cells': [],
                'max_text_length': 0,
                'avg_text_length': 0,
                'text_lengths': []
            }

            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell_analysis['total_cells'] += 1

                    if cell.value is not None:
                        cell_analysis['non_empty_cells'] += 1

                        # 检查数据类型
                        if isinstance(cell.value, str):
                            cell_analysis['text_cells'] += 1
                            text_length = len(cell.value)
                            cell_analysis['text_lengths'].append(text_length)

                            if text_length > cell_analysis['max_text_length']:
                                cell_analysis['max_text_length'] = text_length

                            # 检查长文本（超过500字符）
                            if text_length > 500:
                                cell_analysis['long_text_cells'].append({
                                    'position': f"{chr(64+col)}{row}",
                                    'length': text_length,
                                    'preview': cell.value[:100] + "..." if len(cell.value) > 100 else cell.value
                                })

                            # 检查特殊字符
                            special_chars = re.findall(r'[^\w\s\u4e00-\u9fff\u3400-\u4dbf\u20000-\u2a6df\u2a700-\u2b73f\u2b740-\u2b81f\u2b820-\u2ceaf\uf900-\ufaff\ufe30-\ufe4f,.!?;:()\[\]{}"\'-]', cell.value)
                            if special_chars:
                                cell_analysis['special_char_cells'].append({
                                    'position': f"{chr(64+col)}{row}",
                                    'special_chars': list(set(special_chars)),
                                    'preview': cell.value[:100] + "..." if len(cell.value) > 100 else cell.value
                                })

                            # 检查可能有问题的内容
                            if (text_length > 1000 or
                                len(special_chars) > 5 or
                                '\n' in cell.value or
                                '\r' in cell.value or
                                '\t' in cell.value):
                                cell_analysis['problematic_cells'].append({
                                    'position': f"{chr(64+col)}{row}",
                                    'length': text_length,
                                    'issues': {
                                        'too_long': text_length > 1000,
                                        'many_special_chars': len(special_chars) > 5,
                                        'has_newlines': '\n' in cell.value or '\r' in cell.value,
                                        'has_tabs': '\t' in cell.value
                                    },
                                    'preview': cell.value[:100] + "..." if len(cell.value) > 100 else cell.value
                                })

                        elif isinstance(cell.value, (int, float)):
                            cell_analysis['numeric_cells'] += 1

                        elif hasattr(cell, 'data_type') and cell.data_type == 'f':
                            cell_analysis['formula_cells'] += 1

            # 计算平均文本长度
            if cell_analysis['text_lengths']:
                cell_analysis['avg_text_length'] = sum(cell_analysis['text_lengths']) / len(cell_analysis['text_lengths'])

            sheet_analysis[sheet_name] = cell_analysis

            # 打印当前Sheet的分析结果
            print(f"   非空单元格: {cell_analysis['non_empty_cells']}/{cell_analysis['total_cells']}")
            print(f"   文本单元格: {cell_analysis['text_cells']}")
            print(f"   数值单元格: {cell_analysis['numeric_cells']}")
            print(f"   公式单元格: {cell_analysis['formula_cells']}")
            print(f"   最长文本: {cell_analysis['max_text_length']} 字符")
            print(f"   平均文本长度: {cell_analysis['avg_text_length']:.1f} 字符")
            print(f"   长文本单元格(>500字符): {len(cell_analysis['long_text_cells'])}")
            print(f"   包含特殊字符的单元格: {len(cell_analysis['special_char_cells'])}")
            print(f"   可能有问题的单元格: {len(cell_analysis['problematic_cells'])}")

            # 显示长文本单元格详情
            if cell_analysis['long_text_cells']:
                print(f"\n   📝 长文本单元格详情:")
                for item in cell_analysis['long_text_cells'][:5]:  # 只显示前5个
                    print(f"      {item['position']}: {item['length']}字符 - {item['preview']}")
                if len(cell_analysis['long_text_cells']) > 5:
                    print(f"      ... 还有 {len(cell_analysis['long_text_cells']) - 5} 个长文本单元格")

            # 显示问题单元格详情
            if cell_analysis['problematic_cells']:
                print(f"\n   ⚠️ 可能有问题的单元格:")
                for item in cell_analysis['problematic_cells'][:3]:  # 只显示前3个
                    issues = [k for k, v in item['issues'].items() if v]
                    print(f"      {item['position']}: {item['length']}字符, 问题: {', '.join(issues)}")
                    print(f"         预览: {item['preview']}")
                if len(cell_analysis['problematic_cells']) > 3:
                    print(f"      ... 还有 {len(cell_analysis['problematic_cells']) - 3} 个问题单元格")

        return sheet_analysis

    except Exception as e:
        print(f"❌ 分析文件时出错: {str(e)}")
        return None

def compare_files(original_analysis, result_analysis):
    """对比原始文件和结果文件"""
    print(f"\n{'='*60}")
    print("文件对比分析")
    print(f"{'='*60}")

    if not original_analysis or not result_analysis:
        print("❌ 无法进行对比分析，文件分析失败")
        return

    # 对比Sheet数量
    orig_sheets = set(original_analysis.keys())
    result_sheets = set(result_analysis.keys())

    print(f"📋 Sheet对比:")
    print(f"   原始文件Sheets: {len(orig_sheets)} - {list(orig_sheets)}")
    print(f"   结果文件Sheets: {len(result_sheets)} - {list(result_sheets)}")

    if orig_sheets != result_sheets:
        print(f"   ⚠️ Sheet不匹配!")
        print(f"      原始独有: {orig_sheets - result_sheets}")
        print(f"      结果独有: {result_sheets - orig_sheets}")

    # 对比每个Sheet的内容
    common_sheets = orig_sheets & result_sheets
    for sheet_name in common_sheets:
        print(f"\n📊 Sheet '{sheet_name}' 对比:")
        orig = original_analysis[sheet_name]
        result = result_analysis[sheet_name]

        print(f"   非空单元格: {orig['non_empty_cells']} → {result['non_empty_cells']}")
        print(f"   文本单元格: {orig['text_cells']} → {result['text_cells']}")
        print(f"   最长文本: {orig['max_text_length']} → {result['max_text_length']} 字符")
        print(f"   平均文本长度: {orig['avg_text_length']:.1f} → {result['avg_text_length']:.1f} 字符")
        print(f"   长文本单元格: {len(orig['long_text_cells'])} → {len(result['long_text_cells'])}")
        print(f"   问题单元格: {len(orig['problematic_cells'])} → {len(result['problematic_cells'])}")

        # 分析可能未翻译的问题单元格
        if len(orig['problematic_cells']) > len(result['problematic_cells']):
            print(f"   ✅ 减少了 {len(orig['problematic_cells']) - len(result['problematic_cells'])} 个问题单元格")
        elif len(orig['problematic_cells']) == len(result['problematic_cells']):
            print(f"   ⚠️ 问题单元格数量未变，可能存在翻译困难")

def generate_report(original_analysis, result_analysis):
    """生成详细的分析报告"""
    print(f"\n{'='*60}")
    print("🔍 翻译困难分析报告")
    print(f"{'='*60}")

    print("\n1️⃣ 可能导致翻译卡住的数据特征:")

    # 分析原始文件中的问题
    total_problematic = 0
    total_long_text = 0

    if original_analysis:
        for sheet_name, analysis in original_analysis.items():
            problematic_count = len(analysis['problematic_cells'])
            long_text_count = len(analysis['long_text_cells'])
            total_problematic += problematic_count
            total_long_text += long_text_count

            if problematic_count > 0 or long_text_count > 0:
                print(f"\n   📋 Sheet '{sheet_name}':")
                print(f"      - {problematic_count} 个问题单元格")
                print(f"      - {long_text_count} 个长文本单元格")
                print(f"      - 最长文本: {analysis['max_text_length']} 字符")

                # 显示具体的问题类型
                if analysis['problematic_cells']:
                    issue_types = {}
                    for cell in analysis['problematic_cells']:
                        for issue, has_issue in cell['issues'].items():
                            if has_issue:
                                issue_types[issue] = issue_types.get(issue, 0) + 1

                    print(f"      - 问题类型: {dict(issue_types)}")

    print(f"\n2️⃣ 总体统计:")
    print(f"   - 总问题单元格: {total_problematic}")
    print(f"   - 总长文本单元格: {total_long_text}")

    print(f"\n3️⃣ 可能导致数据库连接问题的数据特征:")
    print("   - 超长文本 (>1000字符) 可能导致数据库写入超时")
    print("   - 包含特殊字符的文本可能导致编码问题")
    print("   - 包含换行符的文本可能导致SQL语句格式错误")
    print("   - 大量并发翻译请求可能导致连接池耗尽")

    print(f"\n4️⃣ 建议的解决方案:")
    print("   - 对超长文本进行分段翻译")
    print("   - 增加特殊字符的转义处理")
    print("   - 优化数据库连接池配置")
    print("   - 添加翻译重试机制")
    print("   - 增加翻译进度的检查点保存")

def main():
    """主函数"""
    print("Excel文件翻译卡住问题分析工具")
    print(f"分析时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # 文件路径
    original_file = "/mnt/d/work/trans_excel/stuck_original.xlsx"
    result_file = "/mnt/d/work/trans_excel/stuck_result.xlsx"

    # 分析原始文件
    print("\n🔍 分析原始文件...")
    original_analysis = analyze_excel_file(original_file)

    # 分析结果文件
    print("\n🔍 分析翻译结果文件...")
    result_analysis = analyze_excel_file(result_file)

    # 对比分析
    compare_files(original_analysis, result_analysis)

    # 生成报告
    generate_report(original_analysis, result_analysis)

    print(f"\n{'='*60}")
    print("✅ 分析完成")
    print(f"{'='*60}")

if __name__ == "__main__":
    main()
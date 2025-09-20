#!/usr/bin/env python3
"""
深度分析卡住的翻译任务 - 找出剩余的6个困难任务
"""
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook
import re
import json
from datetime import datetime
import sys
import os

def analyze_remaining_tasks(original_file, result_file):
    """分析剩余未完成的翻译任务"""
    print(f"\n{'='*80}")
    print("🔍 剩余困难任务深度分析")
    print(f"{'='*80}")

    try:
        # 加载工作簿
        orig_wb = load_workbook(original_file, data_only=True)
        result_wb = load_workbook(result_file, data_only=True)

        remaining_tasks = []

        for sheet_name in orig_wb.sheetnames:
            if sheet_name not in result_wb.sheetnames:
                print(f"❌ 结果文件中缺少Sheet: {sheet_name}")
                continue

            orig_ws = orig_wb[sheet_name]
            result_ws = result_wb[sheet_name]

            print(f"\n📋 分析Sheet: {sheet_name}")
            print(f"   原始尺寸: {orig_ws.max_row} x {orig_ws.max_column}")
            print(f"   结果尺寸: {result_ws.max_row} x {result_ws.max_column}")

            # 对比每个需要翻译的单元格
            for row in range(1, orig_ws.max_row + 1):
                # 检查英文列(B列)
                orig_cell = orig_ws.cell(row=row, column=2)  # B列
                if orig_cell.value and isinstance(orig_cell.value, str):
                    # 检查对应的翻译列是否为空
                    for col in range(3, 6):  # C, D, E列 (葡萄牙语, 泰语, 印尼语)
                        if row <= result_ws.max_row and col <= result_ws.max_column:
                            result_cell = result_ws.cell(row=row, column=col)
                            if not result_cell.value or result_cell.value.strip() == "":
                                # 找到未翻译的单元格
                                lang_names = {3: "Portuguese", 4: "Thai", 5: "Indonesian"}
                                task_info = {
                                    'sheet': sheet_name,
                                    'row': row,
                                    'column': col,
                                    'language': lang_names.get(col, f"Column{col}"),
                                    'original_text': orig_cell.value,
                                    'text_length': len(orig_cell.value),
                                    'issues': analyze_text_issues(orig_cell.value)
                                }
                                remaining_tasks.append(task_info)

        # 分析剩余任务
        print(f"\n📊 剩余未翻译任务统计:")
        print(f"   总计: {len(remaining_tasks)} 个任务")

        # 按语言分组
        by_language = {}
        for task in remaining_tasks:
            lang = task['language']
            if lang not in by_language:
                by_language[lang] = []
            by_language[lang].append(task)

        for lang, tasks in by_language.items():
            print(f"   {lang}: {len(tasks)} 个任务")

        # 分析最困难的任务
        print(f"\n🔥 最困难的任务 (前10个):")
        # 按文本长度和问题数量排序
        sorted_tasks = sorted(remaining_tasks,
                             key=lambda x: (len(x['issues']), x['text_length']),
                             reverse=True)

        for i, task in enumerate(sorted_tasks[:10]):
            print(f"\n   {i+1}. Sheet: {task['sheet']}, 行: {task['row']}, 语言: {task['language']}")
            print(f"      文本长度: {task['text_length']} 字符")
            print(f"      问题类型: {', '.join(task['issues']) if task['issues'] else '无明显问题'}")
            preview = task['original_text'][:100] + "..." if len(task['original_text']) > 100 else task['original_text']
            print(f"      预览: {preview}")

        # 分析问题模式
        print(f"\n📈 问题模式分析:")
        issue_stats = {}
        for task in remaining_tasks:
            for issue in task['issues']:
                issue_stats[issue] = issue_stats.get(issue, 0) + 1

        for issue, count in sorted(issue_stats.items(), key=lambda x: x[1], reverse=True):
            print(f"   {issue}: {count} 次")

        return remaining_tasks

    except Exception as e:
        print(f"❌ 分析时出错: {str(e)}")
        return []

def analyze_text_issues(text):
    """分析文本中可能导致翻译困难的问题"""
    issues = []

    if len(text) > 500:
        issues.append("超长文本")

    if len(text) > 1000:
        issues.append("极长文本")

    # 检查特殊字符
    special_chars = re.findall(r'[^\w\s\u4e00-\u9fff\u3400-\u4dbf\u20000-\u2a6df\u2a700-\u2b73f\u2b740-\u2b81f\u2b820-\u2ceaf\uf900-\ufaff\ufe30-\ufe4f,.!?;:()\[\]{}"\'-]', text)
    if len(special_chars) > 5:
        issues.append("过多特殊字符")

    # 检查换行符
    if '\n' in text or '\r' in text:
        issues.append("包含换行符")

    # 检查制表符
    if '\t' in text:
        issues.append("包含制表符")

    # 检查XML/HTML标签
    if '<' in text and '>' in text:
        issues.append("包含标签")

    # 检查图像引用
    if '<image>' in text:
        issues.append("包含图像引用")

    # 检查样式标签
    if '<style' in text:
        issues.append("包含样式标签")

    # 检查占位符
    if '%s' in text or '%d' in text or '{' in text and '}' in text:
        issues.append("包含占位符")

    # 检查数字和特殊格式
    if re.search(r'\d+\.\s', text):  # 编号列表
        issues.append("包含编号列表")

    return issues

def analyze_database_patterns(remaining_tasks):
    """分析可能导致数据库问题的模式"""
    print(f"\n{'='*80}")
    print("💾 数据库连接问题分析")
    print(f"{'='*80}")

    # 按问题类型分组
    db_risk_tasks = []
    for task in remaining_tasks:
        risk_level = 0
        risk_reasons = []

        # 超长文本风险
        if task['text_length'] > 1000:
            risk_level += 3
            risk_reasons.append("超长文本可能导致数据库写入超时")
        elif task['text_length'] > 500:
            risk_level += 2
            risk_reasons.append("长文本可能增加数据库负载")

        # 特殊字符风险
        if "过多特殊字符" in task['issues']:
            risk_level += 2
            risk_reasons.append("特殊字符可能导致编码问题")

        # 换行符风险
        if "包含换行符" in task['issues']:
            risk_level += 2
            risk_reasons.append("换行符可能导致SQL语句错误")

        # 标签风险
        if "包含标签" in task['issues'] or "包含样式标签" in task['issues']:
            risk_level += 1
            risk_reasons.append("标签内容可能导致解析错误")

        if risk_level > 0:
            task['db_risk_level'] = risk_level
            task['db_risk_reasons'] = risk_reasons
            db_risk_tasks.append(task)

    # 按风险等级排序
    db_risk_tasks.sort(key=lambda x: x['db_risk_level'], reverse=True)

    print(f"📊 数据库风险任务统计:")
    print(f"   高风险任务 (≥4分): {len([t for t in db_risk_tasks if t['db_risk_level'] >= 4])}")
    print(f"   中风险任务 (2-3分): {len([t for t in db_risk_tasks if 2 <= t['db_risk_level'] < 4])}")
    print(f"   低风险任务 (1分): {len([t for t in db_risk_tasks if t['db_risk_level'] == 1])}")

    print(f"\n🔥 高风险任务详情:")
    high_risk_tasks = [t for t in db_risk_tasks if t['db_risk_level'] >= 4]
    for i, task in enumerate(high_risk_tasks[:6]):  # 显示前6个，对应题目中的6个困难任务
        print(f"\n   {i+1}. Sheet: {task['sheet']}, 行: {task['row']}, 语言: {task['language']}")
        print(f"      风险等级: {task['db_risk_level']} 分")
        print(f"      文本长度: {task['text_length']} 字符")
        print(f"      风险原因: {'; '.join(task['db_risk_reasons'])}")
        preview = task['original_text'][:80] + "..." if len(task['original_text']) > 80 else task['original_text']
        print(f"      文本预览: {preview}")

    return db_risk_tasks

def generate_solutions(remaining_tasks, db_risk_tasks):
    """生成针对性的解决方案"""
    print(f"\n{'='*80}")
    print("💡 针对性解决方案")
    print(f"{'='*80}")

    print("\n1️⃣ 即时解决方案:")
    print("   🔧 手动处理高风险任务:")
    high_risk = [t for t in db_risk_tasks if t['db_risk_level'] >= 4][:6]
    for i, task in enumerate(high_risk):
        print(f"      {i+1}. Sheet '{task['sheet']}' 行 {task['row']} - {task['language']}")
        print(f"         建议: 分段翻译，预处理特殊字符")

    print(f"\n2️⃣ 系统优化方案:")
    print("   📚 数据库层面:")
    print("      - 增加连接池大小 (当前可能不足)")
    print("      - 设置更长的查询超时时间")
    print("      - 启用连接重试机制")
    print("      - 优化文本字段的存储引擎")

    print(f"\n   🔄 翻译引擎层面:")
    print("      - 实现文本预处理管道")
    print("      - 添加特殊字符转义功能")
    print("      - 实现长文本分段翻译")
    print("      - 增加翻译失败重试逻辑")

    print(f"\n   📊 监控和恢复:")
    print("      - 实现翻译进度检查点")
    print("      - 添加详细的错误日志")
    print("      - 建立任务队列优先级机制")

def main():
    """主函数"""
    print("Excel翻译卡住问题深度分析")
    print(f"分析时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # 文件路径
    original_file = "/mnt/d/work/trans_excel/stuck_original.xlsx"
    result_file = "/mnt/d/work/trans_excel/stuck_result.xlsx"

    # 分析剩余任务
    remaining_tasks = analyze_remaining_tasks(original_file, result_file)

    if remaining_tasks:
        # 分析数据库风险
        db_risk_tasks = analyze_database_patterns(remaining_tasks)

        # 生成解决方案
        generate_solutions(remaining_tasks, db_risk_tasks)
    else:
        print("✅ 没有发现剩余的未翻译任务")

    print(f"\n{'='*80}")
    print("✅ 深度分析完成")
    print(f"{'='*80}")

if __name__ == "__main__":
    main()
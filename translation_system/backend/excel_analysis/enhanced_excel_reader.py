"""
增强的Excel读取器 - 支持读取单元格颜色和批注信息
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment
from typing import Dict, List, Any, Optional, Tuple
import logging

logger = logging.getLogger(__name__)


class CellMetadata:
    """单元格元数据"""
    def __init__(self):
        self.value = None
        self.comment = None
        self.font_color = None
        self.fill_color = None
        self.font_bold = False
        self.font_italic = False
        self.font_size = None
        self.font_name = None

    def to_dict(self):
        """转换为字典格式"""
        return {
            'value': self.value,
            'comment': self.comment,
            'font_color': self.font_color,
            'fill_color': self.fill_color,
            'font_bold': self.font_bold,
            'font_italic': self.font_italic,
            'font_size': self.font_size,
            'font_name': self.font_name
        }


class EnhancedExcelReader:
    """增强的Excel读取器，支持读取格式和批注"""

    def __init__(self):
        """初始化读取器"""
        self.workbook = None
        self.metadata_cache = {}

    def read_excel_with_metadata(self, file_path: str, sheet_name: Optional[str] = None) -> Tuple[pd.DataFrame, Dict[str, Dict[str, CellMetadata]]]:
        """
        读取Excel文件，包括单元格格式和批注

        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称，如果为None则读取活动工作表

        Returns:
            (DataFrame, 元数据字典)
            元数据字典格式: {sheet_name: {cell_address: CellMetadata}}
        """
        try:
            # 使用openpyxl加载工作簿
            self.workbook = load_workbook(file_path, data_only=True, keep_vba=False)

            # 确定要读取的工作表
            if sheet_name:
                sheets = [sheet_name] if sheet_name in self.workbook.sheetnames else []
            else:
                sheets = self.workbook.sheetnames

            all_data = {}
            all_metadata = {}

            for sheet in sheets:
                ws = self.workbook[sheet]
                sheet_data = []
                sheet_metadata = {}

                # 读取数据和元数据
                for row_idx, row in enumerate(ws.iter_rows(), start=1):
                    row_data = []
                    for col_idx, cell in enumerate(row, start=1):
                        # 创建单元格元数据
                        metadata = CellMetadata()
                        metadata.value = cell.value

                        # 读取批注
                        if cell.comment:
                            metadata.comment = cell.comment.text
                            logger.debug(f"发现批注 [{sheet}!{cell.coordinate}]: {metadata.comment[:50]}...")

                        # 读取字体信息
                        if cell.font:
                            if cell.font.color and cell.font.color.rgb:
                                metadata.font_color = f"#{cell.font.color.rgb}" if isinstance(cell.font.color.rgb, str) else None
                            metadata.font_bold = cell.font.bold or False
                            metadata.font_italic = cell.font.italic or False
                            metadata.font_size = cell.font.size
                            metadata.font_name = cell.font.name

                        # 读取填充色
                        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                            # 过滤掉默认的白色背景
                            if cell.fill.fgColor.rgb not in ['00000000', 'FFFFFFFF', None]:
                                metadata.fill_color = f"#{cell.fill.fgColor.rgb}" if isinstance(cell.fill.fgColor.rgb, str) else None
                                logger.debug(f"发现填充色 [{sheet}!{cell.coordinate}]: {metadata.fill_color}")

                        # 保存元数据
                        cell_address = cell.coordinate
                        if metadata.comment or metadata.fill_color or metadata.font_color:
                            sheet_metadata[cell_address] = metadata

                        row_data.append(cell.value)

                    sheet_data.append(row_data)

                # 转换为DataFrame
                if sheet_data:
                    # 使用第一行作为列名
                    if len(sheet_data) > 1:
                        df = pd.DataFrame(sheet_data[1:], columns=sheet_data[0])
                    else:
                        df = pd.DataFrame(sheet_data)

                    all_data[sheet] = df
                    all_metadata[sheet] = sheet_metadata

                    # 统计信息
                    comments_count = sum(1 for m in sheet_metadata.values() if m.comment)
                    colors_count = sum(1 for m in sheet_metadata.values() if m.fill_color or m.font_color)
                    logger.info(f"Sheet '{sheet}': 发现 {comments_count} 个批注, {colors_count} 个带颜色的单元格")

            # 如果只有一个sheet，返回单个DataFrame
            if len(all_data) == 1:
                sheet_name = list(all_data.keys())[0]
                return all_data[sheet_name], {sheet_name: all_metadata[sheet_name]}
            else:
                # 返回第一个sheet的数据，但保留所有元数据
                first_sheet = list(all_data.keys())[0]
                return all_data[first_sheet], all_metadata

        except Exception as e:
            logger.error(f"读取Excel文件失败: {e}")
            # 降级到普通的pandas读取
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            return df, {}

    def get_all_sheets_with_metadata(self, file_path: str) -> Dict[str, Tuple[pd.DataFrame, Dict[str, CellMetadata]]]:
        """
        读取所有工作表的数据和元数据

        Args:
            file_path: Excel文件路径

        Returns:
            {sheet_name: (DataFrame, {cell_address: CellMetadata})}
        """
        try:
            self.workbook = load_workbook(file_path, data_only=True, keep_vba=False)
            all_sheets = {}

            for sheet_name in self.workbook.sheetnames:
                df, metadata = self.read_excel_with_metadata(file_path, sheet_name)
                all_sheets[sheet_name] = (df, metadata.get(sheet_name, {}))

            return all_sheets

        except Exception as e:
            logger.error(f"读取所有工作表失败: {e}")
            return {}

    def write_excel_with_metadata(self, df: pd.DataFrame, output_path: str,
                                 metadata: Dict[str, CellMetadata],
                                 sheet_name: str = 'Sheet1'):
        """
        写入Excel文件，保留格式和批注

        Args:
            df: 数据DataFrame
            output_path: 输出文件路径
            metadata: 元数据字典 {cell_address: CellMetadata}
            sheet_name: 工作表名称
        """
        try:
            # 先用pandas写入数据
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                # 获取工作簿和工作表
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]

                # 应用元数据
                for cell_address, cell_metadata in metadata.items():
                    try:
                        cell = worksheet[cell_address]

                        # 添加批注
                        if cell_metadata.comment:
                            comment = Comment(cell_metadata.comment, "Translation System")
                            cell.comment = comment

                        # 设置字体颜色
                        if cell_metadata.font_color:
                            from openpyxl.styles import Font
                            color_hex = cell_metadata.font_color.replace('#', '')
                            cell.font = Font(
                                color=color_hex,
                                bold=cell_metadata.font_bold,
                                italic=cell_metadata.font_italic,
                                size=cell_metadata.font_size,
                                name=cell_metadata.font_name
                            )

                        # 设置填充色
                        if cell_metadata.fill_color:
                            from openpyxl.styles import PatternFill
                            color_hex = cell_metadata.fill_color.replace('#', '')
                            cell.fill = PatternFill(
                                start_color=color_hex,
                                end_color=color_hex,
                                fill_type='solid'
                            )

                    except Exception as e:
                        logger.warning(f"应用元数据到单元格 {cell_address} 失败: {e}")

            logger.info(f"✅ Excel文件已保存（包含格式和批注）: {output_path}")

        except Exception as e:
            logger.error(f"写入Excel文件失败: {e}")
            # 降级到普通的pandas写入
            df.to_excel(output_path, sheet_name=sheet_name, index=False)

    def extract_comments_as_context(self, metadata: Dict[str, CellMetadata]) -> Dict[str, str]:
        """
        提取批注作为翻译上下文

        Args:
            metadata: 元数据字典

        Returns:
            {cell_address: comment_text}
        """
        comments = {}
        for cell_address, cell_metadata in metadata.items():
            if cell_metadata.comment:
                comments[cell_address] = cell_metadata.comment
        return comments

    def get_colored_cells(self, metadata: Dict[str, CellMetadata]) -> Dict[str, Dict[str, str]]:
        """
        获取所有带颜色的单元格

        Args:
            metadata: 元数据字典

        Returns:
            {cell_address: {'font_color': color, 'fill_color': color}}
        """
        colored_cells = {}
        for cell_address, cell_metadata in metadata.items():
            if cell_metadata.font_color or cell_metadata.fill_color:
                colored_cells[cell_address] = {
                    'font_color': cell_metadata.font_color,
                    'fill_color': cell_metadata.fill_color,
                    'value': cell_metadata.value
                }
        return colored_cells


# 使用示例
if __name__ == "__main__":
    # 设置日志
    logging.basicConfig(level=logging.DEBUG)

    reader = EnhancedExcelReader()

    # 读取Excel文件
    df, metadata = reader.read_excel_with_metadata("test.xlsx")

    # 提取批注
    comments = reader.extract_comments_as_context(metadata.get('Sheet1', {}))
    print(f"批注: {comments}")

    # 获取带颜色的单元格
    colored = reader.get_colored_cells(metadata.get('Sheet1', {}))
    print(f"带颜色的单元格: {colored}")

    # 写入新文件（保留格式）
    reader.write_excel_with_metadata(df, "output.xlsx", metadata.get('Sheet1', {}))
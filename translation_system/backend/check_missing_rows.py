#!/usr/bin/env python3
"""
检查缺失的行和翻译任务
"""
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook
import re
import json
from datetime import datetime
import sys
import os

def detailed_row_analysis(original_file, result_file):
    """详细分析每一行的翻译状态"""
    print(f"\n{'='*80}")
    print("🔍 逐行翻译状态分析")
    print(f"{'='*80}")

    try:
        # 加载工作簿
        orig_wb = load_workbook(original_file, data_only=True)
        result_wb = load_workbook(result_file, data_only=True)

        all_missing_tasks = []

        for sheet_name in orig_wb.sheetnames:
            print(f"\n📋 分析Sheet: {sheet_name}")

            orig_ws = orig_wb[sheet_name]

            if sheet_name not in result_wb.sheetnames:
                print(f"❌ 结果文件中完全缺少Sheet: {sheet_name}")
                continue

            result_ws = result_wb[sheet_name]

            print(f"   原始: {orig_ws.max_row} 行 x {orig_ws.max_column} 列")
            print(f"   结果: {result_ws.max_row} 行 x {result_ws.max_column} 列")

            # 统计各种状态
            original_texts = 0
            translated_rows = 0
            missing_rows = 0
            partial_translations = 0
            problematic_texts = []

            for row in range(1, orig_ws.max_row + 1):
                # 检查B列（英文原文）
                orig_cell_b = orig_ws.cell(row=row, column=2)
                if orig_cell_b.value and isinstance(orig_cell_b.value, str) and orig_cell_b.value.strip():
                    original_texts += 1

                    # 检查这一行在结果文件中是否存在
                    if row > result_ws.max_row:
                        missing_rows += 1
                        all_missing_tasks.append({
                            'sheet': sheet_name,
                            'row': row,
                            'type': 'missing_row',
                            'original_text': orig_cell_b.value,
                            'text_length': len(orig_cell_b.value),
                            'issues': analyze_text_complexity(orig_cell_b.value)
                        })
                        continue

                    # 检查翻译状态
                    row_translation_status = {
                        'Portuguese': False,  # C列
                        'Thai': False,        # D列
                        'Indonesian': False   # E列
                    }

                    # 检查C列（葡萄牙语）
                    if result_ws.max_column >= 3:
                        result_cell_c = result_ws.cell(row=row, column=3)
                        if result_cell_c.value and str(result_cell_c.value).strip():
                            row_translation_status['Portuguese'] = True

                    # 检查D列（泰语）
                    if result_ws.max_column >= 4:
                        result_cell_d = result_ws.cell(row=row, column=4)
                        if result_cell_d.value and str(result_cell_d.value).strip():
                            row_translation_status['Thai'] = True

                    # 检查E列（印尼语）
                    if result_ws.max_column >= 5:
                        result_cell_e = result_ws.cell(row=row, column=5)
                        if result_cell_e.value and str(result_cell_e.value).strip():
                            row_translation_status['Indonesian'] = True

                    # 统计翻译状态
                    translations_completed = sum(row_translation_status.values())

                    if translations_completed == 3:
                        translated_rows += 1
                    elif translations_completed > 0:
                        partial_translations += 1
                        # 记录部分翻译的任务
                        for lang, completed in row_translation_status.items():
                            if not completed:
                                all_missing_tasks.append({
                                    'sheet': sheet_name,
                                    'row': row,
                                    'type': 'missing_translation',
                                    'language': lang,
                                    'original_text': orig_cell_b.value,
                                    'text_length': len(orig_cell_b.value),
                                    'issues': analyze_text_complexity(orig_cell_b.value)
                                })
                    else:
                        # 完全没有翻译
                        for lang in ['Portuguese', 'Thai', 'Indonesian']:
                            all_missing_tasks.append({
                                'sheet': sheet_name,
                                'row': row,
                                'type': 'no_translation',
                                'language': lang,
                                'original_text': orig_cell_b.value,
                                'text_length': len(orig_cell_b.value),
                                'issues': analyze_text_complexity(orig_cell_b.value)
                            })

            print(f"   📊 统计:")
            print(f"      原始文本数: {original_texts}")
            print(f"      完全翻译行: {translated_rows}")
            print(f"      部分翻译行: {partial_translations}")
            print(f"      缺失行数: {missing_rows}")
            print(f"      翻译完成率: {(translated_rows/original_texts*100 if original_texts > 0 else 0):.1f}%")

        # 分析缺失任务
        print(f"\n📈 缺失任务分析:")
        print(f"   总缺失任务: {len(all_missing_tasks)}")

        # 按类型分组
        by_type = {}
        for task in all_missing_tasks:
            task_type = task['type']
            if task_type not in by_type:
                by_type[task_type] = []
            by_type[task_type].append(task)

        for task_type, tasks in by_type.items():
            print(f"   {task_type}: {len(tasks)} 个")

        # 按语言分组
        by_language = {}
        for task in all_missing_tasks:
            if 'language' in task:
                lang = task['language']
                if lang not in by_language:
                    by_language[lang] = []
                by_language[lang].append(task)

        print(f"\n📊 按语言分组:")
        for lang, tasks in by_language.items():
            print(f"   {lang}: {len(tasks)} 个任务")

        return all_missing_tasks

    except Exception as e:
        print(f"❌ 分析时出错: {str(e)}")
        import traceback
        traceback.print_exc()
        return []

def analyze_text_complexity(text):
    """分析文本复杂度"""
    issues = []

    if len(text) > 300:
        issues.append("长文本")
    if len(text) > 600:
        issues.append("超长文本")

    # 特殊字符
    special_chars = re.findall(r'[<>]', text)
    if len(special_chars) > 3:
        issues.append("HTML标签")

    if '\n' in text:
        issues.append("换行符")
    if '\t' in text:
        issues.append("制表符")
    if '%s' in text or '%d' in text:
        issues.append("占位符")
    if re.search(r'\d+\.', text):
        issues.append("编号列表")

    return issues

def find_hardest_tasks(missing_tasks):
    """找出最困难的任务"""
    print(f"\n{'='*80}")
    print("🔥 最困难任务识别")
    print(f"{'='*80}")

    # 计算困难度分数
    for task in missing_tasks:
        difficulty_score = 0

        # 文本长度得分
        length = task['text_length']
        if length > 600:
            difficulty_score += 4
        elif length > 300:
            difficulty_score += 2
        elif length > 100:
            difficulty_score += 1

        # 问题类型得分
        for issue in task['issues']:
            if issue == "超长文本":
                difficulty_score += 3
            elif issue == "HTML标签":
                difficulty_score += 2
            elif issue in ["换行符", "编号列表"]:
                difficulty_score += 2
            else:
                difficulty_score += 1

        task['difficulty_score'] = difficulty_score

    # 排序找出最困难的
    hardest_tasks = sorted(missing_tasks,
                          key=lambda x: x['difficulty_score'],
                          reverse=True)

    print(f"🎯 最困难的6个任务:")
    for i, task in enumerate(hardest_tasks[:6]):
        print(f"\n   {i+1}. Sheet: {task['sheet']}, 行: {task['row']}")
        print(f"      语言: {task.get('language', 'All')}")
        print(f"      类型: {task['type']}")
        print(f"      困难度: {task['difficulty_score']} 分")
        print(f"      文本长度: {task['text_length']} 字符")
        print(f"      问题: {', '.join(task['issues']) if task['issues'] else '无特殊问题'}")
        preview = task['original_text'][:80] + "..." if len(task['original_text']) > 80 else task['original_text']
        print(f"      预览: {preview}")

    return hardest_tasks[:6]

def main():
    """主函数"""
    print("Excel翻译缺失任务详细分析")
    print(f"分析时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # 文件路径
    original_file = "/mnt/d/work/trans_excel/stuck_original.xlsx"
    result_file = "/mnt/d/work/trans_excel/stuck_result.xlsx"

    # 详细分析
    missing_tasks = detailed_row_analysis(original_file, result_file)

    if missing_tasks:
        # 找出最困难的任务
        hardest_tasks = find_hardest_tasks(missing_tasks)

        print(f"\n{'='*80}")
        print("💡 解决建议")
        print(f"{'='*80}")
        print("1. 优先处理困难度最高的6个任务")
        print("2. 对超长文本进行分段处理")
        print("3. 预处理HTML标签和特殊字符")
        print("4. 增加数据库连接的重试机制")
        print("5. 实现翻译进度的断点续传")
    else:
        print("✅ 所有翻译任务已完成")

    print(f"\n{'='*80}")
    print("✅ 分析完成")
    print(f"{'='*80}")

if __name__ == "__main__":
    main()
"""Download API for translated Excel files."""

import os
import logging
from pathlib import Path
from typing import Optional
from fastapi import APIRouter, HTTPException, Response
from fastapi.responses import FileResponse, JSONResponse

from services.export.excel_exporter import excel_exporter
from utils.pipeline_session_manager import pipeline_session_manager


logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api", tags=["download"])


@router.get("/download/{session_id}")
async def download_translated_excel(session_id: str):
    """
    Download translated Excel file.

    Args:
        session_id: Session identifier

    Returns:
        FileResponse: Translated Excel file
    """
    try:
        # Validate session exists
        session = pipeline_session_manager.get_session(session_id)
        if not session:
            raise HTTPException(
                status_code=404,
                detail=f"Session {session_id} not found"
            )

        # ✅ FIX: Check if execution is completed
        if session.stage.value != "completed":
            raise HTTPException(
                status_code=400,
                detail=f"Cannot download: execution not completed (current stage: {session.stage.value})"
            )

        # ✅ FIX: Verify output_state exists
        output_state = pipeline_session_manager.get_output_state(session_id)
        output_state_timestamp = session.metadata.get('output_state_timestamp')

        if not output_state and not output_state_timestamp:
            raise HTTPException(
                status_code=400,
                detail="Cannot download: no output data available"
            )

        # ✅ FIX: Check if cached export is valid
        export_info = excel_exporter.get_export_info(session_id)
        cache_is_valid = False

        if export_info['has_export'] and export_info['file_exists']:
            # Compare cached export timestamp with output_state timestamp
            export_timestamp = export_info.get('export_timestamp')

            if export_timestamp and output_state_timestamp:
                from datetime import datetime
                try:
                    export_time = datetime.fromisoformat(export_timestamp.replace('Z', '+00:00'))
                    output_time = datetime.fromisoformat(output_state_timestamp.replace('Z', '+00:00'))

                    # Cache is valid if export was created AFTER output_state
                    if export_time >= output_time:
                        cache_is_valid = True
                        logger.info(f"Cache valid: export_time={export_time} >= output_time={output_time}")
                    else:
                        logger.warning(f"Cache stale: export_time={export_time} < output_time={output_time}, will regenerate")
                except Exception as e:
                    logger.warning(f"Failed to compare timestamps: {e}, will regenerate")
            else:
                # If no timestamps, use cache (backward compatibility)
                cache_is_valid = True

        # Return cached file if valid
        if cache_is_valid:
            file_path = export_info['exported_file']
            filename = Path(file_path).name

            logger.info(f"Returning cached export file: {filename}")

            return FileResponse(
                path=file_path,
                filename=filename,
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

        # ✅ Clear stale cache if exists
        if export_info['has_export'] and export_info['file_exists']:
            try:
                os.remove(export_info['exported_file'])
                logger.info(f"Removed stale cache file: {export_info['exported_file']}")
            except Exception as e:
                logger.warning(f"Failed to remove stale cache: {e}")

        # Generate new export
        try:
            exported_file = await excel_exporter.export_final_excel(session_id)
            filename = Path(exported_file).name

            logger.info(f"Generated new export file: {filename}")

            return FileResponse(
                path=exported_file,
                filename=filename,
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

        except ValueError as e:
            raise HTTPException(
                status_code=400,
                detail=f"Cannot export file: {str(e)}"
            )
        except Exception as e:
            logger.error(f"Failed to export file for session {session_id}: {e}")
            raise HTTPException(
                status_code=500,
                detail="Failed to generate export file"
            )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Unexpected error in download endpoint: {e}")
        raise HTTPException(
            status_code=500,
            detail="Internal server error"
        )


@router.get("/download/{session_id}/input")
async def download_input_excel(session_id: str):
    """
    Download input Excel file (before transformation).

    This exports the input_state (ExcelDataFrame) back to Excel format.
    Useful for Stage 1 to verify what was uploaded/inherited.

    Args:
        session_id: Session identifier

    Returns:
        FileResponse: Input Excel file
    """
    try:
        # Validate session exists
        session = pipeline_session_manager.get_session(session_id)
        if not session:
            raise HTTPException(
                status_code=404,
                detail=f"Session {session_id} not found"
            )

        # Check if session has input_state
        input_state = pipeline_session_manager.get_input_state(session_id)
        if not input_state:
            # Try to load from file
            input_file_path = session.metadata.get('input_file_path')
            if input_file_path and os.path.exists(input_file_path):
                # Return the original file directly
                filename = Path(input_file_path).name
                return FileResponse(
                    path=input_file_path,
                    filename=f"input_{session.session_id[:8]}_{filename}",
                    media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            else:
                raise HTTPException(
                    status_code=404,
                    detail=f"Session {session_id} has no input state"
                )

        # Export input_state (ExcelDataFrame) to Excel
        import tempfile
        from openpyxl import Workbook

        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            output_path = tmp_file.name

        # Create workbook from ExcelDataFrame
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        for sheet_name, df in input_state.sheets.items():
            ws = wb.create_sheet(title=sheet_name)

            # Write headers
            for col_idx, col_name in enumerate(df.columns):
                ws.cell(row=1, column=col_idx + 1, value=col_name)

            # Write data
            for row_idx in range(len(df)):
                for col_idx in range(len(df.columns)):
                    value = df.iloc[row_idx, col_idx]
                    ws.cell(row=row_idx + 2, column=col_idx + 1, value=value)

        wb.save(output_path)

        return FileResponse(
            path=output_path,
            filename=f"input_state_{session.session_id[:8]}.xlsx",
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to export input for session {session_id}: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to export input Excel: {str(e)}"
        )


@router.get("/download/{session_id}/summary")
async def download_task_summary(session_id: str):
    """
    Download task summary Excel file.

    Args:
        session_id: Session identifier

    Returns:
        FileResponse: Task summary Excel file
    """
    try:
        # Validate session exists
        session = pipeline_session_manager.get_session(session_id)
        if not session:
            raise HTTPException(
                status_code=404,
                detail=f"Session {session_id} not found"
            )

        # Generate task summary
        try:
            summary_file = await excel_exporter.export_task_summary(session_id)
            filename = Path(summary_file).name

            logger.info(f"Generated task summary file: {filename}")

            return FileResponse(
                path=summary_file,
                filename=filename,
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

        except ValueError as e:
            raise HTTPException(
                status_code=400,
                detail=f"Cannot generate summary: {str(e)}"
            )
        except Exception as e:
            logger.error(f"Failed to generate summary for session {session_id}: {e}")
            raise HTTPException(
                status_code=500,
                detail="Failed to generate summary file"
            )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Unexpected error in summary endpoint: {e}")
        raise HTTPException(
            status_code=500,
            detail="Internal server error"
        )


@router.get("/download/{session_id}/info")
async def get_download_info(session_id: str):
    """
    Get download information for a session.

    Args:
        session_id: Session identifier

    Returns:
        JSON response with download information including execution status
    """
    try:
        # Validate session exists
        session = pipeline_session_manager.get_session(session_id)
        if not session:
            raise HTTPException(
                status_code=404,
                detail=f"Session {session_id} not found"
            )

        # Get task dataframe and statistics
        task_manager = pipeline_session_manager.get_tasks(session_id)

        if task_manager is None or task_manager.df is None or task_manager.df.empty:
            # No tasks - need to split tasks first
            return JSONResponse(content={
                "session_id": session_id,
                "status": "no_tasks",
                "message": "任务尚未拆分，无法下载",
                "ready_for_download": False,
                "can_download": False,
                "execution_status": None,
                "task_statistics": {},
                "export_info": {
                    "has_export": False,
                    "file_exists": False
                }
            })

        # Calculate statistics from task_df
        task_df = task_manager.df
        total = len(task_df)
        completed = len(task_df[task_df['status'] == 'completed'])
        failed = len(task_df[task_df['status'] == 'failed'])
        processing = len(task_df[task_df['status'] == 'processing'])
        pending = len(task_df[task_df['status'] == 'pending'])

        task_stats = {
            'total': total,
            'completed': completed,
            'failed': failed,
            'processing': processing,
            'pending': pending
        }

        # Check execution stage using TransformationStage
        stage = session.stage
        ready_for_download = False
        status = "unknown"
        message = ""

        # Map TransformationStage to execution status
        if stage.value == "completed":
            status = "completed"
            message = "翻译已完成，可以下载结果"
            ready_for_download = True
        elif stage.value == "executing":
            status = "executing"
            completed = task_stats.get('completed', 0)
            total = task_stats.get('total', 0)
            message = f"翻译进行中 ({completed}/{total})，请等待完成后下载"
        elif stage.value == "split_complete":
            status = "not_started"
            message = "任务已拆分但尚未执行，请先开始翻译"
        elif stage.value == "input_loaded":
            status = "not_split"
            message = "数据已加载但未拆分任务，请先拆分"
        elif stage.value == "created":
            status = "not_ready"
            message = "Session刚创建，未准备好"
        elif stage.value == "failed":
            status = "failed"
            message = "翻译失败"
            ready_for_download = session.output_state is not None  # 可能有部分结果
        else:
            status = stage.value
            message = f"状态: {stage.value}"

        execution_status = {
            "stage": stage.value,
            "ready_for_download": ready_for_download,
            "has_output": session.output_state is not None
        }

        # Get export information
        export_info = excel_exporter.get_export_info(session_id)

        # Determine if can download
        # Can download if:
        # 1. Execution is completed (ready_for_download=True), OR
        # 2. Already has exported file
        can_download = ready_for_download or export_info['has_export']

        # Combine information
        download_info = {
            "session_id": session_id,
            "status": status,
            "message": message,
            "ready_for_download": ready_for_download,
            "can_download": can_download,
            "execution_status": execution_status,
            "task_statistics": task_stats,
            "export_info": export_info
        }

        return JSONResponse(content=download_info)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get download info for session {session_id}: {e}")
        raise HTTPException(
            status_code=500,
            detail="Failed to retrieve download information"
        )


@router.delete("/download/{session_id}/files")
async def cleanup_session_files(session_id: str):
    """
    Clean up exported files for a session.

    Args:
        session_id: Session identifier

    Returns:
        JSON response with cleanup results
    """
    try:
        # Validate session exists
        session = pipeline_session_manager.get_session(session_id)
        if not session:
            raise HTTPException(
                status_code=404,
                detail=f"Session {session_id} not found"
            )

        cleanup_count = 0
        cleanup_errors = []

        # Get export info
        export_info = excel_exporter.get_export_info(session_id)

        if export_info['has_export'] and export_info['file_exists']:
            try:
                os.remove(export_info['exported_file'])
                cleanup_count += 1

                # Clear metadata
                pipeline_session_manager.set_metadata(session_id, 'exported_file', None)
                pipeline_session_manager.set_metadata(session_id, 'export_timestamp', None)

                logger.info(f"Cleaned up export file for session {session_id}")

            except Exception as e:
                error_msg = f"Failed to remove export file: {str(e)}"
                cleanup_errors.append(error_msg)
                logger.error(error_msg)

        # Clean up any other files in output directory matching session
        try:
            output_dir = excel_exporter.output_dir
            for file_path in output_dir.glob(f"*{session_id[:8]}*.xlsx"):
                try:
                    file_path.unlink()
                    cleanup_count += 1
                except Exception as e:
                    error_msg = f"Failed to remove file {file_path}: {str(e)}"
                    cleanup_errors.append(error_msg)

        except Exception as e:
            error_msg = f"Failed to scan output directory: {str(e)}"
            cleanup_errors.append(error_msg)

        return JSONResponse(content={
            "session_id": session_id,
            "files_removed": cleanup_count,
            "errors": cleanup_errors,
            "success": len(cleanup_errors) == 0
        })

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to cleanup files for session {session_id}: {e}")
        raise HTTPException(
            status_code=500,
            detail="Failed to cleanup session files"
        )
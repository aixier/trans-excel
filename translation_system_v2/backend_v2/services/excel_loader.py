"""Excel loader service - DataFrame Pipeline Architecture.

核心设计：
- 加载时将颜色和注释直接存入 DataFrame 的 color_* 和 comment_* 列
- 不再使用单独的 color_map 和 comment_map
- 保持 DataFrame Pipeline 的格式一致性
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from typing import Optional, Dict, Any
import uuid
from pathlib import Path

from models.excel_dataframe import ExcelDataFrame


class ExcelLoader:
    """Load Excel files with color and comment extraction."""

    @staticmethod
    def load_excel(file_path: str) -> ExcelDataFrame:
        """
        Load Excel file with all metadata.

        返回的 DataFrame 包含：
        - 数据列：key, CH, EN, TH, ...
        - 颜色列：color_CH, color_EN, color_TH, ...
        - 注释列：comment_CH, comment_EN, comment_TH, ...
        """
        excel_df = ExcelDataFrame()
        excel_df.filename = Path(file_path).name
        excel_df.excel_id = str(uuid.uuid4())

        # Load workbook for metadata extraction
        wb = openpyxl.load_workbook(file_path, data_only=False)

        # Load all sheets
        excel_data = pd.read_excel(file_path, sheet_name=None)

        for sheet_name, df in excel_data.items():
            # Get worksheet
            ws = wb[sheet_name]

            # Get column names from DataFrame
            columns = df.columns.tolist()

            # Create color and comment columns for each data column
            for col_name in columns:
                df[f'color_{col_name}'] = None  # Initialize color column
                df[f'comment_{col_name}'] = None  # Initialize comment column

            # Extract colors and comments from openpyxl
            # Note: openpyxl row/col are 1-based, pandas DataFrame are 0-based
            for row_idx, row in enumerate(ws.iter_rows(min_row=2)):  # Skip header row
                if row_idx >= len(df):  # Safety check
                    break

                for col_idx, cell in enumerate(row):
                    if col_idx >= len(columns):  # Safety check
                        break

                    col_name = columns[col_idx]

                    # Extract color
                    if cell.fill and cell.fill.patternType:
                        fill = cell.fill
                        if isinstance(fill.fgColor.rgb, str) and fill.fgColor.rgb:
                            color = f"#{fill.fgColor.rgb}"
                            if color != "#00000000":  # Ignore transparent
                                df.at[row_idx, f'color_{col_name}'] = color

                    # Extract comment
                    if cell.comment:
                        df.at[row_idx, f'comment_{col_name}'] = cell.comment.text

            # Add the enhanced DataFrame to ExcelDataFrame
            excel_df.add_sheet(sheet_name, df)

        wb.close()
        return excel_df

    @staticmethod
    def save_excel(excel_df: ExcelDataFrame, file_path: str) -> None:
        """
        Save Excel file with colors and comments.

        从 DataFrame 的 color_* 和 comment_* 列读取元数据并应用到 Excel。
        """
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Write all sheets (only data columns, not color_* and comment_*)
            for sheet_name, df in excel_df.sheets.items():
                # Get data columns (exclude color_* and comment_*)
                data_cols = excel_df.get_data_columns(sheet_name)
                df_to_write = df[data_cols]

                df_to_write.to_excel(writer, sheet_name=sheet_name, index=False)

            # Get workbook and apply colors/comments
            wb = writer.book

            for sheet_name in excel_df.get_sheet_names():
                ws = wb[sheet_name]
                df = excel_df.get_sheet(sheet_name)
                data_cols = excel_df.get_data_columns(sheet_name)

                # Apply colors and comments by reading from DataFrame columns
                for row_idx in range(len(df)):
                    for col_idx, col_name in enumerate(data_cols):
                        # Apply color
                        color = df.at[row_idx, f'color_{col_name}']
                        if pd.notna(color) and color and color.startswith('#'):
                            cell = ws.cell(row=row_idx+2, column=col_idx+1)  # +2 for header, +1 for 1-based
                            fill = PatternFill(
                                start_color=color[1:],  # Remove '#'
                                end_color=color[1:],
                                fill_type='solid'
                            )
                            cell.fill = fill

                        # Apply comment
                        comment_text = df.at[row_idx, f'comment_{col_name}']
                        if pd.notna(comment_text) and comment_text:
                            cell = ws.cell(row=row_idx+2, column=col_idx+1)
                            cell.comment = openpyxl.comments.Comment(str(comment_text), "System")

    @staticmethod
    def validate_excel(file_path: str) -> Dict[str, Any]:
        """Validate Excel file format."""
        try:
            # Try to load the file
            excel_data = pd.read_excel(file_path, sheet_name=None, nrows=5)

            return {
                'valid': True,
                'sheets': list(excel_data.keys()),
                'sheet_count': len(excel_data)
            }
        except Exception as e:
            return {
                'valid': False,
                'error': str(e)
            }

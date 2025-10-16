"""Excel export optimization service - DataFrame Pipeline Architecture.

核心改动：
- 从 DataFrame 的 color_* 和 comment_* 列读取元数据
- 导出时只处理数据列，不处理元数据列
- 不再使用 excel_df.get_cell_color() 和 excel_df.get_cell_comment()
"""

import os
import logging
from pathlib import Path
from typing import Dict, Any, Optional, List
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from datetime import datetime

from models.excel_dataframe import ExcelDataFrame
from models.task_dataframe import TaskDataFrameManager, TaskStatus
from utils.pipeline_session_manager import pipeline_session_manager


logger = logging.getLogger(__name__)


class ExcelExporter:
    """Excel exporter with format preservation and optimization."""

    def __init__(self, output_dir: str = None):
        """
        Initialize ExcelExporter.

        Args:
            output_dir: Output directory for exported files
        """
        self.output_dir = Path(output_dir) if output_dir else Path("./output")
        self.output_dir.mkdir(exist_ok=True)
        self.logger = logging.getLogger(self.__class__.__name__)

    async def export_final_excel(self, session_id: str) -> str:
        """
        Export final Excel file with translation results.

        Args:
            session_id: Session identifier

        Returns:
            Path to the exported Excel file
        """
        try:
            # Get session first
            session = pipeline_session_manager.get_session(session_id)
            if not session:
                raise ValueError(f"Session {session_id} not found")

            # Get session data using correct pipeline API
            # Try output_state first, fallback to input_state if not available
            excel_df = pipeline_session_manager.get_output_state(session_id)
            if not excel_df:
                self.logger.warning(f"No output_state for session {session_id}, using input_state")
                excel_df = pipeline_session_manager.get_input_state(session_id)

            # ✅ FIX: If still no excel_df (pickle deserialization failed), reload from file
            if not excel_df:
                input_file_path = pipeline_session_manager.get_metadata(session_id, 'input_file_path')
                if input_file_path and os.path.exists(input_file_path):
                    from services.excel_loader import ExcelLoader
                    self.logger.warning(f"input_state unavailable, reloading Excel from file: {input_file_path}")
                    loader = ExcelLoader()
                    excel_df = loader.load_excel(input_file_path)
                    self.logger.info(f"✅ Reloaded Excel from file: {len(excel_df.sheets)} sheets")

            task_manager = pipeline_session_manager.get_tasks(session_id)

            if not excel_df or not task_manager:
                raise ValueError(f"Session {session_id} data not found or incomplete")

            # ✅ FIX: Force reload task_manager from file to get latest translations
            task_file_path = pipeline_session_manager.get_metadata(session_id, 'task_file_path')
            if task_file_path and os.path.exists(task_file_path):
                import pandas as pd
                from models.task_dataframe import TaskDataFrameManager

                self.logger.info(f"Reloading task_manager from file for latest data: {task_file_path}")
                task_manager = TaskDataFrameManager()
                task_manager.df = pd.read_parquet(task_file_path)
                self.logger.info(f"Reloaded {len(task_manager.df)} tasks from file")

            # Create output filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_filename = Path(excel_df.filename).stem
            output_filename = f"{base_filename}_translated_{timestamp}.xlsx"
            output_path = self.output_dir / output_filename

            self.logger.info(f"Exporting Excel file to: {output_path}")

            # Create new workbook
            wb = Workbook()

            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # Process each sheet
            for sheet_name, df in excel_df.sheets.items():
                ws = wb.create_sheet(title=sheet_name)

                # Write data with translations
                await self._write_sheet_with_translations(
                    ws, df, excel_df, task_manager, sheet_name
                )

                # Apply formatting
                self._apply_sheet_formatting(ws, df, excel_df, sheet_name)

            # Save workbook
            wb.save(output_path)

            # Store export metadata with ISO format timestamp for comparison
            export_timestamp_iso = datetime.now().isoformat()
            pipeline_session_manager.set_metadata(
                session_id, 'exported_file', str(output_path)
            )
            pipeline_session_manager.set_metadata(
                session_id, 'export_timestamp', export_timestamp_iso
            )

            self.logger.info(f"Successfully exported Excel file: {output_path} at {export_timestamp_iso}")
            return str(output_path)

        except Exception as e:
            self.logger.error(f"Failed to export Excel for session {session_id}: {e}")
            raise

    async def _write_sheet_with_translations(
        self,
        worksheet,
        df: pd.DataFrame,
        excel_df: ExcelDataFrame,
        task_manager: TaskDataFrameManager,
        sheet_name: str
    ):
        """Write sheet data with translation results."""

        # Get data columns only (exclude color_* and comment_*)
        data_cols = excel_df.get_data_columns(sheet_name)

        # Get completed tasks for this sheet
        if task_manager.df is not None:
            sheet_tasks = task_manager.df[
                (task_manager.df['sheet_name'] == sheet_name) &
                (task_manager.df['status'] == TaskStatus.COMPLETED)
            ]
        else:
            sheet_tasks = pd.DataFrame()

        # Create translation map: (row, col) -> translated_text
        # Sort tasks to ensure caps tasks override normal tasks
        if not sheet_tasks.empty:
            sheet_tasks = sheet_tasks.sort_values('task_type', ascending=True)  # caps < normal alphabetically

        translation_map = {}
        for _, task in sheet_tasks.iterrows():
            row_idx = int(task['row_idx'])
            col_name = task['col_name']  # Use col_name instead of col_idx
            # Caps tasks will be processed last and override normal tasks
            translation_map[(row_idx, col_name)] = task['result']

        # ✅ Write column headers first (row 1) - only data columns
        for excel_col_idx, col_name in enumerate(data_cols):
            cell = worksheet.cell(row=1, column=excel_col_idx + 1)
            cell.value = col_name

        # Write data row by row (starting from row 2)
        for row_idx in range(len(df)):
            for excel_col_idx, col_name in enumerate(data_cols):
                # Get original value
                original_value = df.at[row_idx, col_name]

                # Use translation if available, otherwise use original
                if (row_idx, col_name) in translation_map:
                    cell_value = translation_map[(row_idx, col_name)]
                else:
                    cell_value = original_value

                # Write to Excel (1-based indexing, +1 for header row)
                excel_row = row_idx + 2  # +2 to account for header row
                excel_col = excel_col_idx + 1
                cell = worksheet.cell(row=excel_row, column=excel_col)
                cell.value = cell_value

                # Add comment if task was translated AND original value was not empty
                if (row_idx, col_name) in translation_map:
                    # Only add comment if original_value is not NaN/None (re-translation case)
                    if pd.notna(original_value) and str(original_value).strip():
                        # Read original comment from DataFrame
                        original_comment = df.at[row_idx, f'comment_{col_name}']
                        if pd.notna(original_comment):
                            original_comment = str(original_comment)
                        else:
                            original_comment = None

                        translation_comment = f"原文: {original_value}"

                        if original_comment:
                            final_comment = f"{original_comment}\n{translation_comment}"
                        else:
                            final_comment = translation_comment

                        cell.comment = Comment(final_comment, "TranslationSystem")

    def _apply_sheet_formatting(
        self,
        worksheet,
        df: pd.DataFrame,
        excel_df: ExcelDataFrame,
        sheet_name: str
    ):
        """Apply formatting to worksheet."""

        # Get data columns only
        data_cols = excel_df.get_data_columns(sheet_name)

        # Get completed tasks for this sheet to identify translated cells
        task_manager = None
        translation_cells = set()

        try:
            # This is called after _write_sheet_with_translations
            # We need to rebuild translation_cells set for formatting
            # Get from session via excel_df.excel_id
            sessions = pipeline_session_manager._sessions
            for sid, sess in sessions.items():
                if sess.excel_df and sess.excel_df.excel_id == excel_df.excel_id:
                    task_manager = sess.task_manager
                    break

            if task_manager and task_manager.df is not None:
                sheet_tasks = task_manager.df[
                    (task_manager.df['sheet_name'] == sheet_name) &
                    (task_manager.df['status'] == TaskStatus.COMPLETED)
                ]
                for _, task in sheet_tasks.iterrows():
                    translation_cells.add((int(task['row_idx']), task['col_name']))
        except Exception as e:
            logger.warning(f"Could not load translation cells for formatting: {e}")

        # Create named styles
        translated_style = NamedStyle(name="translated")
        translated_style.fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
        translated_style.font = Font(italic=True)

        # Apply cell colors and formatting (skip header row)
        for row_idx in range(len(df)):
            for excel_col_idx, col_name in enumerate(data_cols):
                excel_row = row_idx + 2  # +2 to account for header row
                excel_col = excel_col_idx + 1
                cell = worksheet.cell(row=excel_row, column=excel_col)

                # Apply original color if exists (read from DataFrame)
                original_color = df.at[row_idx, f'color_{col_name}']

                if pd.notna(original_color) and original_color:
                    # Remove '#' if present
                    color = str(original_color).lstrip('#')
                    fill = PatternFill(
                        start_color=color, end_color=color, fill_type="solid"
                    )
                    cell.fill = fill

                # ✅ Mark translated cells (both first-time and re-translation)
                if (row_idx, col_name) in translation_cells:
                    # Add gray background to translated cells if no original color
                    current_fill = cell.fill
                    if not current_fill or current_fill.start_color.rgb == '00000000':
                        cell.fill = translated_style.fill
                    # Italic font for all translated cells
                    cell.font = Font(italic=True)

        # Auto-adjust column widths
        self._auto_adjust_columns(worksheet, data_cols, df)

    def _auto_adjust_columns(self, worksheet, data_cols: List[str], df: pd.DataFrame):
        """Auto-adjust column widths based on content."""

        for excel_col_idx, col_name in enumerate(data_cols):
            column_letter = get_column_letter(excel_col_idx + 1)

            # Calculate max width for column (including header)
            max_width = len(str(col_name))  # Start with header length

            for row_idx in range(len(df)):
                cell_value = df.at[row_idx, col_name]
                if pd.notna(cell_value):
                    cell_length = len(str(cell_value))
                    max_width = max(max_width, cell_length)

            # Set width (with reasonable limits)
            width = min(max_width + 2, 50)  # Max 50 characters
            worksheet.column_dimensions[column_letter].width = width

    async def export_task_summary(self, session_id: str) -> str:
        """
        Export task summary Excel file.

        Args:
            session_id: Session identifier

        Returns:
            Path to the summary Excel file
        """
        try:
            task_manager = pipeline_session_manager.get_tasks(session_id)

            if not task_manager or task_manager.df is None:
                raise ValueError(f"No tasks found for session {session_id}")

            # Create output filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            summary_filename = f"task_summary_{session_id[:8]}_{timestamp}.xlsx"
            summary_path = self.output_dir / summary_filename

            # Export task DataFrame to Excel
            with pd.ExcelWriter(summary_path, engine='openpyxl') as writer:
                # Main tasks sheet
                task_manager.df.to_excel(writer, sheet_name='Tasks', index=False)

                # Statistics sheet
                stats = task_manager.get_statistics()
                stats_df = pd.DataFrame([stats])
                stats_df.to_excel(writer, sheet_name='Statistics', index=False)

                # By status sheet
                if stats['by_status']:
                    status_df = pd.DataFrame(
                        list(stats['by_status'].items()),
                        columns=['Status', 'Count']
                    )
                    status_df.to_excel(writer, sheet_name='By_Status', index=False)

            self.logger.info(f"Exported task summary to: {summary_path}")
            return str(summary_path)

        except Exception as e:
            self.logger.error(f"Failed to export task summary: {e}")
            raise

    def get_export_info(self, session_id: str) -> Dict[str, Any]:
        """
        Get export information for a session.

        Args:
            session_id: Session identifier

        Returns:
            Export information
        """
        exported_file = pipeline_session_manager.get_metadata(session_id, 'exported_file')
        export_timestamp = pipeline_session_manager.get_metadata(session_id, 'export_timestamp')

        info = {
            'has_export': exported_file is not None,
            'exported_file': exported_file,
            'export_timestamp': export_timestamp,
            'file_exists': False,
            'file_size': None
        }

        if exported_file and os.path.exists(exported_file):
            info['file_exists'] = True
            info['file_size'] = os.path.getsize(exported_file)

        return info

    async def cleanup_old_exports(self, days_old: int = 7):
        """
        Clean up old exported files.

        Args:
            days_old: Remove files older than this many days
        """
        try:
            if not self.output_dir.exists():
                return

            current_time = datetime.now()
            cleanup_count = 0

            for file_path in self.output_dir.glob("*.xlsx"):
                # Check file age
                file_time = datetime.fromtimestamp(file_path.stat().st_mtime)
                age_days = (current_time - file_time).days

                if age_days > days_old:
                    try:
                        file_path.unlink()
                        cleanup_count += 1
                        self.logger.debug(f"Removed old export file: {file_path}")
                    except Exception as e:
                        self.logger.warning(f"Failed to remove file {file_path}: {e}")

            if cleanup_count > 0:
                self.logger.info(f"Cleaned up {cleanup_count} old export files")

        except Exception as e:
            self.logger.error(f"Failed to cleanup old exports: {e}")


# Global exporter instance
excel_exporter = ExcelExporter()

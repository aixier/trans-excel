#!/usr/bin/env python3
"""
测试数据生成脚本
用于生成AI翻译系统的测试Excel文件
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
import random
import os

class TestDataGenerator:
    """测试数据生成器"""

    def __init__(self, output_dir="test_data"):
        """初始化生成器"""
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

        # 示例游戏文本
        self.sample_texts = {
            'welcome': '欢迎来到游戏世界',
            'start_game': '开始游戏',
            'continue': '继续',
            'settings': '设置',
            'exit': '退出',
            'player_name': '玩家 %s',
            'gold_gained': '你获得了 {%d} 金币',
            'level_up': '<font color="gold">恭喜升级!</font>',
            'quest_complete': '任务完成\n获得奖励',
            'item_description': '这是一把传说中的剑，据说曾经属于勇者',
            'dialog_npc': '年轻的冒险者，你愿意帮助我们村庄吗？',
            'skill_desc': '<b>火球术</b>\n造成 {damage} 点火焰伤害',
            'achievement': '成就解锁：<H1>屠龙勇士</H1>',
        }

        # 颜色定义
        self.yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

    def generate_basic_test(self):
        """生成基础功能测试文件"""
        data = {
            'KEY': [],
            'CH': [],
            'EN': [],
            'JP': [],
            'KR': [],
            'FR': [],
            'DE': [],
            'ES': [],
            'PT': [],
            'RU': [],
            'TR': []
        }

        # 生成100行测试数据
        for i in range(100):
            key = f"text_{i:03d}"
            data['KEY'].append(key)

            # 随机选择中文文本
            sample_key = random.choice(list(self.sample_texts.keys()))
            ch_text = self.sample_texts[sample_key]

            # 添加一些变化
            if i % 10 == 0:
                ch_text = f"<font size={12 + i % 5}>{ch_text}</font>"
            elif i % 15 == 0:
                ch_text = ch_text + f" {{num_{i}}}"
            elif i % 20 == 0:
                ch_text = f"{ch_text}\n第二行文本"

            data['CH'].append(ch_text)

            # 英文列（部分填充，模拟增量翻译场景）
            if i < 80:
                data['EN'].append(f"English text {i}")
            else:
                data['EN'].append("")

            # 其他语言列（部分留空，测试新语言翻译）
            for lang in ['JP', 'KR', 'FR', 'DE', 'ES', 'PT', 'RU', 'TR']:
                if i < 50 and lang in ['JP', 'KR']:
                    data[lang].append(f"{lang} text {i}")
                else:
                    data[lang].append("")

        # 创建DataFrame并保存
        df = pd.DataFrame(data)
        output_file = os.path.join(self.output_dir, "basic_test.xlsx")

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='GameText', index=False)

        print(f"✅ 生成基础测试文件: {output_file}")
        return output_file

    def generate_color_test(self):
        """生成颜色标记测试文件"""
        data = {
            'KEY': [],
            'CH': [],
            'EN': [],
            'JP': [],
            'KR': []
        }

        # 生成30行数据
        for i in range(30):
            data['KEY'].append(f"color_test_{i:02d}")
            data['CH'].append(f"测试文本 {i}")
            data['EN'].append(f"Test text {i}" if i < 20 else "")
            data['JP'].append(f"テストテキスト {i}" if i < 15 else "")
            data['KR'].append(f"테스트 텍스트 {i}" if i < 15 else "")

        # 创建Excel
        df = pd.DataFrame(data)
        output_file = os.path.join(self.output_dir, "color_test.xlsx")

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='ColorTest', index=False)

            # 获取工作表
            workbook = writer.book
            worksheet = writer.sheets['ColorTest']

            # 添加黄色标记（行5-14，表示需要重新翻译）
            for row in range(6, 16):  # Excel行号从1开始，加上标题行
                for col in range(2, 6):  # B-E列
                    cell = worksheet.cell(row=row, column=col)
                    cell.fill = self.yellow_fill

            # 添加蓝色标记（行15-24，表示需要缩短翻译）
            for row in range(16, 26):
                for col in range(2, 6):
                    cell = worksheet.cell(row=row, column=col)
                    cell.fill = self.blue_fill

        print(f"✅ 生成颜色标记测试文件: {output_file}")
        return output_file

    def generate_comment_test(self):
        """生成带批注的测试文件"""
        data = {
            'KEY': [],
            'CH': [],
            'EN': [],
            'JP': []
        }

        comments_config = [
            ('npc_dialog_01', '村长的对话', '这是一位德高望重的老人，语气要庄重'),
            ('ui_button_start', '开始游戏', 'UI按钮，尽量简短'),
            ('item_sword', '传说之剑', '这是游戏中最强的武器，翻译要有史诗感'),
            ('skill_fire', '火球术', '技能名称，保持简洁有力'),
            ('quest_title', '拯救公主', '主线任务标题，要吸引人'),
            ('achievement_01', '初出茅庐', '成就名称，4个字为佳'),
            ('shop_welcome', '欢迎光临', 'Simon: 缩短这句话的翻译'),
            ('battle_victory', '战斗胜利！', '要有胜利的喜悦感'),
            ('game_over', '游戏结束', '不要太沮丧，鼓励玩家再试一次'),
            ('loading_tip', '小提示：记得及时保存游戏', '加载界面提示，保持轻松友好'),
        ]

        for key, text, comment in comments_config:
            data['KEY'].append(key)
            data['CH'].append(text)
            data['EN'].append("")
            data['JP'].append("")

        # 创建Excel
        df = pd.DataFrame(data)
        output_file = os.path.join(self.output_dir, "comment_test.xlsx")

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='CommentTest', index=False)

            # 获取工作表
            worksheet = writer.sheets['CommentTest']

            # 添加批注
            for i, (_, _, comment_text) in enumerate(comments_config, start=2):
                cell = worksheet.cell(row=i, column=2)  # B列（中文）
                comment = Comment(comment_text, 'GameDesigner')
                cell.comment = comment

        print(f"✅ 生成批注测试文件: {output_file}")
        return output_file

    def generate_multi_sheet_test(self):
        """生成多Sheet测试文件"""
        output_file = os.path.join(self.output_dir, "multi_sheet_test.xlsx")

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # UI表（UI_开头，触发UI翻译模式）
            ui_data = {
                'KEY': ['btn_ok', 'btn_cancel', 'btn_confirm', 'menu_settings', 'menu_exit'],
                'CH': ['确定', '取消', '确认', '设置', '退出'],
                'EN': ['', '', '', '', ''],
                'JP': ['', '', '', '', '']
            }
            pd.DataFrame(ui_data).to_excel(writer, sheet_name='UI_Buttons', index=False)

            # 对话表
            dialog_data = {
                'KEY': ['npc_01', 'npc_02', 'player_01'],
                'CH': [
                    '勇者啊，请帮助我们消灭恶龙！',
                    '这把剑是我祖传的宝物，现在送给你。',
                    '我一定会完成任务的！'
                ],
                'EN': ['', '', ''],
                'JP': ['', '', '']
            }
            pd.DataFrame(dialog_data).to_excel(writer, sheet_name='Dialog_NPC', index=False)

            # 物品描述表
            item_data = {
                'KEY': ['sword_01', 'potion_01', 'armor_01'],
                'CH': [
                    '一把锋利的铁剑，适合初学者使用',
                    '恢复50点生命值的药水',
                    '提供基础防御的皮甲'
                ],
                'EN': ['', '', ''],
                'JP': ['', '', '']
            }
            pd.DataFrame(item_data).to_excel(writer, sheet_name='Item_Description', index=False)

            # 游戏背景表（特殊表）
            game_desc = {
                'KEY': ['game_type', 'world_view', 'target_market'],
                'Description': [
                    '日式RPG游戏',
                    '西方魔幻世界观，包含龙、精灵、矮人等种族',
                    '目标发行地区：东南亚、日韩、欧美'
                ]
            }
            pd.DataFrame(game_desc).to_excel(writer, sheet_name='Game_Description', index=False)

        print(f"✅ 生成多Sheet测试文件: {output_file}")
        return output_file

    def generate_format_test(self):
        """生成格式化标记测试文件"""
        data = {
            'KEY': [],
            'CH': [],
            'EN': [],
            'JP': []
        }

        format_tests = [
            ('format_01', '玩家<b>%s</b>加入了游戏'),
            ('format_02', '你获得了<font color="gold">{%d}</font>金币'),
            ('format_03', '任务进度：{current}/{total}'),
            ('format_04', '<H1>第一章</H1>\n<H2>觉醒</H2>'),
            ('format_05', '攻击力+{atk_bonus}%\n防御力+{def_bonus}%'),
            ('format_06', '使用技能<skill>{skill_name}</skill>'),
            ('format_07', '倒计时：{time}秒'),
            ('format_08', '<font size="12" color="#FF0000">警告！</font>'),
            ('format_09', '等级 {level} - 经验 {exp}/{max_exp}'),
            ('format_10', '第{chapter}章 - {title}'),
            ('format_11', '<i>这是一段斜体文字</i>'),
            ('format_12', '<u>带下划线的重要提示</u>'),
            ('format_13', '列表：\n• 项目1\n• 项目2\n• 项目3'),
            ('format_14', '复合格式：<b>玩家{name}</b>在<font color="red">危险区域</font>'),
            ('format_15', 'HTML实体：&lt;标签&gt; &amp; 符号'),
        ]

        for key, text in format_tests:
            data['KEY'].append(key)
            data['CH'].append(text)
            data['EN'].append('')
            data['JP'].append('')

        # 保存文件
        df = pd.DataFrame(data)
        output_file = os.path.join(self.output_dir, "format_test.xlsx")
        df.to_excel(output_file, index=False, sheet_name='FormatTest')

        print(f"✅ 生成格式化标记测试文件: {output_file}")
        return output_file

    def generate_performance_test(self):
        """生成性能测试文件（大文件）"""
        data = {
            'KEY': [],
            'CH': [],
            'EN': [],
            'JP': [],
            'KR': [],
            'FR': [],
            'DE': [],
            'ES': [],
            'PT': [],
            'RU': [],
            'TR': [],
            'AR': [],
            'TH': [],
            'VN': [],
            'ID': []
        }

        # 生成5000行数据
        text_templates = [
            "游戏文本示例 {}",
            "这是第{}个任务的描述",
            "物品{}的详细说明",
            "技能{}：造成{damage}点伤害",
            "第{}章 - 标题文本",
            "NPC对话 {} - 很长的一段对话内容，用来测试性能" * 5,
        ]

        for i in range(5000):
            data['KEY'].append(f"perf_test_{i:05d}")

            # 随机选择模板
            template = random.choice(text_templates)
            ch_text = template.format(i)

            # 20%的概率添加格式
            if random.random() < 0.2:
                ch_text = f"<font>{ch_text}</font>"

            data['CH'].append(ch_text)

            # 英文列80%填充
            if random.random() < 0.8:
                data['EN'].append(f"English text {i}")
            else:
                data['EN'].append("")

            # 其他语言随机填充
            for lang in ['JP', 'KR', 'FR', 'DE', 'ES', 'PT', 'RU', 'TR', 'AR', 'TH', 'VN', 'ID']:
                if random.random() < 0.3:  # 30%已翻译
                    data[lang].append(f"{lang} text {i}")
                else:
                    data[lang].append("")

        # 创建多个Sheet
        output_file = os.path.join(self.output_dir, "performance_test.xlsx")

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # 将5000行分成5个Sheet，每个1000行
            for sheet_num in range(5):
                start_idx = sheet_num * 1000
                end_idx = (sheet_num + 1) * 1000

                sheet_data = {}
                for col in data.keys():
                    sheet_data[col] = data[col][start_idx:end_idx]

                df = pd.DataFrame(sheet_data)
                df.to_excel(writer, sheet_name=f'Sheet{sheet_num + 1}', index=False)

        print(f"✅ 生成性能测试文件: {output_file} (5000行, 5个Sheet)")
        return output_file

    def generate_terms_file(self):
        """生成术语表文件"""
        terms = {
            "game_terms": {
                "CH": {
                    "法师": {
                        "EN": "Mage",
                        "JP": "メイジ",
                        "KR": "마법사",
                        "FR": "Mage",
                        "DE": "Magier"
                    },
                    "战士": {
                        "EN": "Warrior",
                        "JP": "戦士",
                        "KR": "전사",
                        "FR": "Guerrier",
                        "DE": "Krieger"
                    },
                    "盗贼": {
                        "EN": "Rogue",
                        "JP": "ローグ",
                        "KR": "도적",
                        "FR": "Voleur",
                        "DE": "Schurke"
                    },
                    "金币": {
                        "EN": "Gold",
                        "JP": "ゴールド",
                        "KR": "골드",
                        "FR": "Or",
                        "DE": "Gold"
                    },
                    "经验值": {
                        "EN": "EXP",
                        "JP": "経験値",
                        "KR": "경험치",
                        "FR": "EXP",
                        "DE": "EXP"
                    },
                    "等级": {
                        "EN": "Level",
                        "JP": "レベル",
                        "KR": "레벨",
                        "FR": "Niveau",
                        "DE": "Stufe"
                    },
                    "任务": {
                        "EN": "Quest",
                        "JP": "クエスト",
                        "KR": "퀘스트",
                        "FR": "Quête",
                        "DE": "Quest"
                    },
                    "装备": {
                        "EN": "Equipment",
                        "JP": "装備",
                        "KR": "장비",
                        "FR": "Équipement",
                        "DE": "Ausrüstung"
                    },
                    "技能": {
                        "EN": "Skill",
                        "JP": "スキル",
                        "KR": "스킬",
                        "FR": "Compétence",
                        "DE": "Fähigkeit"
                    },
                    "公会": {
                        "EN": "Guild",
                        "JP": "ギルド",
                        "KR": "길드",
                        "FR": "Guilde",
                        "DE": "Gilde"
                    }
                }
            },
            "ui_terms": {
                "CH": {
                    "确定": {"EN": "OK", "JP": "OK", "KR": "확인"},
                    "取消": {"EN": "Cancel", "JP": "キャンセル", "KR": "취소"},
                    "返回": {"EN": "Back", "JP": "戻る", "KR": "뒤로"},
                    "设置": {"EN": "Settings", "JP": "設定", "KR": "설정"},
                    "开始": {"EN": "Start", "JP": "スタート", "KR": "시작"}
                }
            }
        }

        # 保存为JSON
        import json
        output_file = os.path.join(self.output_dir, "terms.json")
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(terms, f, ensure_ascii=False, indent=2)

        print(f"✅ 生成术语表文件: {output_file}")

        # 同时生成Excel格式的术语表
        terms_data = {
            'CH': [],
            'EN': [],
            'JP': [],
            'KR': [],
            'FR': [],
            'DE': [],
            'Category': []
        }

        for category, category_terms in terms.items():
            for ch_term, translations in category_terms['CH'].items():
                terms_data['CH'].append(ch_term)
                terms_data['EN'].append(translations.get('EN', ''))
                terms_data['JP'].append(translations.get('JP', ''))
                terms_data['KR'].append(translations.get('KR', ''))
                terms_data['FR'].append(translations.get('FR', ''))
                terms_data['DE'].append(translations.get('DE', ''))
                terms_data['Category'].append(category)

        df = pd.DataFrame(terms_data)
        excel_file = os.path.join(self.output_dir, "terms.xlsx")
        df.to_excel(excel_file, index=False, sheet_name='Terms')

        print(f"✅ 生成术语表Excel: {excel_file}")
        return output_file, excel_file

    def generate_all(self):
        """生成所有测试文件"""
        print("=" * 50)
        print("开始生成测试数据...")
        print("=" * 50)

        files = []
        files.append(self.generate_basic_test())
        files.append(self.generate_color_test())
        files.append(self.generate_comment_test())
        files.append(self.generate_multi_sheet_test())
        files.append(self.generate_format_test())
        files.append(self.generate_performance_test())
        terms_files = self.generate_terms_file()
        files.extend(terms_files)

        print("=" * 50)
        print(f"✅ 所有测试文件已生成完成！")
        print(f"📁 输出目录: {os.path.abspath(self.output_dir)}")
        print("=" * 50)

        # 生成文件清单
        manifest_file = os.path.join(self.output_dir, "test_files_manifest.txt")
        with open(manifest_file, 'w', encoding='utf-8') as f:
            f.write("测试文件清单\n")
            f.write("=" * 50 + "\n\n")
            for file in files:
                if file:
                    f.write(f"- {os.path.basename(file)}\n")
                    # 获取文件大小
                    size = os.path.getsize(file) / 1024  # KB
                    f.write(f"  大小: {size:.2f} KB\n\n")

        print(f"📝 文件清单已保存: {manifest_file}")
        return files


def main():
    """主函数"""
    generator = TestDataGenerator()

    # 生成所有测试文件
    generator.generate_all()

    print("\n使用说明:")
    print("-" * 30)
    print("1. 基础功能测试: basic_test.xlsx")
    print("2. 颜色标记测试: color_test.xlsx")
    print("3. 批注测试: comment_test.xlsx")
    print("4. 多Sheet测试: multi_sheet_test.xlsx")
    print("5. 格式化测试: format_test.xlsx")
    print("6. 性能测试: performance_test.xlsx")
    print("7. 术语表: terms.json / terms.xlsx")


if __name__ == "__main__":
    main()